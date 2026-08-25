import os
import stat
import json
import shutil
import base64
import re
from pathlib import Path

# =====================================================================
# 1) CONFIGURAÇÃO
# =====================================================================
BASE_DIR       = Path(__file__).parent.resolve()
IMG_DIR        = BASE_DIR / "imagens_produtos"
# --- planilha: detecta automaticamente o .xlsx da pasta -------------
def _achar_planilha():
    cands = [f for f in BASE_DIR.glob("*.xls*")
             if not f.name.startswith("~$")]
    if not cands:
        return BASE_DIR / "dados_produtos.xlsx"
    # prioridade: "stock" > "dados_produtos" > mais recente
    for chave in ("stock", "estoque", "dados_produtos"):
        for f in cands:
            if chave in f.name.lower():
                return f
    return max(cands, key=lambda f: f.stat().st_mtime)

XLSX_PATH      = _achar_planilha()
LOGO_PATH      = BASE_DIR / "logo_glab.png"
OUTPUT_DIR     = BASE_DIR / "site_glab"
OUTPUT_HTML    = OUTPUT_DIR / "index.html"
OUTPUT_IMG_DIR = OUTPUT_DIR / "imagens_produtos"

WHATSAPP_NUM   = "17746222523"   # ajuste para o número real
FRETE_GRATIS   = 2000.0            # frete grátis acima deste valor
BRINDE_LIMITE  = 1000.0            # a partir daqui, ganha bacteriostatic water por item

# ---------------------------------------------------------------------
# RASTREIO DE CLIENTES (Google Sheets via Apps Script)
# Cole aqui a URL do Web App publicado (arquivo apps_script_glab.gs).
# Deixe "" para desativar o rastreio.
# ---------------------------------------------------------------------
ANALYTICS_URL  = ""

# =====================================================================
# 2) BASE DE PRODUTOS (mesmos 31 peptídeos do site original)
# =====================================================================
CATEGORY_COLORS = {
    "Metabolismo":   "#ff3b30",
    "Hormônios":     "#ff5252",
    "Recuperação":   "#e53935",
    "Estética":      "#ff1744",
    "Imunidade":     "#d32f2f",
    "Emagrecimento": "#ff6b6b",
    "Cognitivo":     "#c62828",
    "Longevidade":   "#b71c1c",
    "Sexual":        "#ef5350",
    "Acessório":     "#9e9e9e",
}

PRODUTOS_PADRAO = [
    # (nome, espec, preco_orig, preco, promo_pct, cat, icon, disponivel, info)
    ("AOD 9604", "5 mg", 220, 220, 0, "Metabolismo", "🔥", True, "Análogo Lipolítico do hGH: foca no isolamento das propriedades de queima de gordura do GH sem induzir efeitos hiperglicêmicos."),
    ("HGH FRAGMENT 176-191", "5 mg", 240, 240, 0, "Metabolismo", "🔥", True, "Modulador de Lipídios: parte terminal do GH responsável pela quebra de gordura."),
    ("MOTS-C", "10 mg", 380, 380, 0, "Metabolismo", "🔥", True, "Peptídeo derivado da mitocôndria: regulador do metabolismo sistêmico via AMPK."),
    ("SLU PP 332", "10 mg", 520, 520, 0, "Metabolismo", "🔥", True, "Agonista Pan-ERR (Pílula do Exercício): aumenta biogênese mitocondrial."),
    ("TESAMORELIN", "5 mg", 340, 340, 0, "Metabolismo", "🔥", True, "Único aprovado para reduzir gordura visceral abdominal severa."),
    ("CJC-1295", "5 mg", 260, 260, 0, "Hormônios", "💉", True, "Secretagogo de GH de longa duração — aumenta GH e IGF-1."),
    ("IPAMORELIN", "5 mg", 240, 240, 0, "Hormônios", "💉", True, "Agonista de grelina seletivo — libera GH sem elevar cortisol."),
    ("CJC-1295 + IPAMORELIN", "10 mg", 420, 420, 0, "Hormônios", "💉", True, "Sinergia GHRH + GHRP: mimetiza liberação fisiológica natural de GH."),
    ("IGF-1 LR3", "1 mg", 480, 480, 0, "Hormônios", "💉", True, "Análogo de IGF-1 de meia-vida longa — hiperplasia muscular."),
    ("SERMORELIN", "5 mg", 220, 220, 0, "Hormônios", "💉", True, "Mimetiza o GHRH natural — melhora sono profundo e vitalidade."),
    ("BPC-157", "5 mg", 260, 260, 0, "Recuperação", "🩹", True, "Pentadecapeptídeo gástrico — acelera cicatrização de tendões, ligamentos e músculos."),
    ("TB-500", "5 mg", 320, 320, 0, "Recuperação", "🩹", True, "Timosina Beta-4 sintética — reparo de tecidos e angiogênese."),
    ("TB-500 + BPC", "10 mg", 520, 520, 0, "Recuperação", "🩹", True, "Protocolo de reparo total — padrão-ouro para lesões atléticas graves."),
    ("ARA 290", "5 mg", 380, 380, 0, "Recuperação", "🩹", True, "Derivado da eritropoietina — dor neuropática e regeneração nervosa."),
    ("KLOW", "10 mg", 560, 560, 0, "Recuperação", "🩹", True, "Quarteto de reparo profundo (GHK+BPC+TB+KPV)."),
    ("GHK-CU", "50 mg", 340, 340, 0, "Estética", "✨", True, "Complexo peptídeo-cobre — remodelação do DNA e síntese de colágeno."),
    ("GLOW", "10 mg", 460, 460, 0, "Estética", "✨", True, "Bioestimulação dérmica (GHK-Cu + BPC + TB) — rejuvenescimento cutâneo."),
    ("KPV", "5 mg", 260, 260, 0, "Imunidade", "🛡️", True, "Tripeptídeo anti-inflamatório — inibe NF-κB."),
    ("TIRZEPATIDE", "10 mg", 780, 780, 0, "Emagrecimento", "⚖️", True, "Agonista dual GIP/GLP-1 — supera a semaglutida na perda de peso."),
    ("RETATRUTIDE", "10 mg", 940, 940, 0, "Emagrecimento", "⚖️", True, "Agonista triplo (GIP/GLP-1/GCGR) — perdas de peso >24%."),
    ("SEMAGLUTIDE", "5 mg", 620, 620, 0, "Emagrecimento", "⚖️", True, "Agonista de GLP-1 — retarda esvaziamento gástrico."),
    ("SELANK", "5 mg", 280, 280, 0, "Cognitivo", "🧠", True, "Ansiolítico regulador — modula serotonina e norepinefrina."),
    ("SEMAX", "5 mg", 300, 300, 0, "Cognitivo", "🧠", True, "Nootrópico neuroprotetor — eleva BDNF e NGF."),
    ("PINEALON", "5 mg", 320, 320, 0, "Cognitivo", "🧠", True, "Bioregulador de cadeia curta — restaura ritmo circadiano."),
    ("DSIP", "5 mg", 280, 280, 0, "Cognitivo", "🧠", True, "Indutor de sono delta — sincroniza ritmos biológicos."),
    ("OXYTOCIN", "5 mg", 240, 240, 0, "Cognitivo", "🧠", False, "Neuromodulador social — regula confiança e ansiedade social."),
    ("NAD+", "500 mg", 460, 460, 0, "Longevidade", "⏳", True, "Coenzima de vitalidade — reparação do DNA e sirtuínas."),
    ("EPITHALON", "10 mg", 360, 360, 0, "Longevidade", "⏳", True, "Ativador da telomerase — extensão da vida celular."),
    ("SS-31", "5 mg", 420, 420, 0, "Longevidade", "⏳", True, "Protetor de cardiolipina — restaura produção de ATP."),
    ("PT-141", "10 mg", 260, 260, 0, "Sexual", "❤️", True, "Tratamento de disfunção sexual — atua no SNC."),
    ("BACTERIOSTATIC WATER", "30 ml", 45, 45, 0, "Acessório", "💧", True, "Água bacteriostática 0,9% álcool benzílico — solvente para reconstituição."),
]

# Emoji específico por produto (chave = nome normalizado por _norm_nome).
# Representa a função principal estudada de cada peptídeo.
ICONES_PRODUTO = {
    "AOD 9604": "🔥",
    "HGH FRAGMENT 176 191": "🔥",
    "MOTS C": "⚡",
    "SLU PP": "🏃",
    "SLU PP 332": "🏃",
    "TESAMORELIN": "🔥",
    "CJC 1295": "💉",
    "IPAMORELIN": "💤",
    "CJC 1295 IPAMORELIN": "💉",
    "IGF 1 LR3": "💪",
    "SERMORELIN": "🌙",
    "BPC 157": "🩹",
    "TB 500": "🦵",
    "TB 500 BPC": "🩹",
    "ARA 290": "⚡",
    "KLOW": "🧩",
    "GHK CU": "✨",
    "GLOW": "💎",
    "KPV": "🛡️",
    "TIRZEPATIDE": "⚖️",
    "RETATRUTIDE": "⚖️",
    "SEMAGLUTIDE": "⚖️",
    "SELANK": "🧘",
    "SEMAX": "🧠",
    "PINEALON": "🕰️",
    "DSIP": "🧠",
    "OXYTOCIN": "❤️",
    "NAD": "🔋",
    "EPITHALON": "⏳",
    "SS 31": "🫀",
    "PT 141": "❤️‍🔥",
    "BACTERIOSTATIC WATER": "💧",
    "CASE": "📦",
    "CBL 514": "🔥",
    "BPC": "🩹",
}



# =====================================================================
# 3) LEITURA DO EXCEL (disponibilidade + descontos + preços)
# =====================================================================
COL_ALIASES = {
    "produto":   ["produto", "nome", "peptideo", "peptídeo", "item", "descricao", "descrição"],
    "volume":    ["volume", "qtd", "quantidade", "espec", "tamanho"],
    "medida":    ["medida", "unidade", "un", "und"],
    "estoque":   ["estoque", "disponivel", "disponível", "disponibilidade", "status"],
    "preco":     ["preco (r$)", "preço (r$)", "preco", "preço", "valor", "valor (r$)", "preco r$", "preço r$"],
    "promocao":  ["promocao", "promoção", "promo", "desconto", "promo_pct", "promoção (%)", "promocao (%)"],
}


def _norm_header(v):
    return " ".join(str(v or "").strip().lower().split())


def _norm_nome(v):
    """Normaliza nome de produto para casar planilha x base (ignora acentos, hífens, espaços)."""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(v or "")).encode("ascii", "ignore").decode()
    s = s.upper()
    for ch in "-+/,.()":
        s = s.replace(ch, " ")
    return " ".join(s.split())


def _parse_preco(v):
    """Aceita 400, 400.0, 'R$ 400,00', '400,00'."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).upper().replace("R$", "").replace(" ", "").replace("\xa0", "")
    s = s.replace(".", "").replace(",", ".") if "," in s else s
    try:
        return float(s)
    except ValueError:
        return None


def _parse_promo(v):
    """Aceita 0.1, 10, '10%', '10 %', vazio. Retorna fração (0.10)."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        p = float(v)
    else:
        s = str(v).replace("%", "").replace(",", ".").strip()
        if not s:
            return None
        try:
            p = float(s)
        except ValueError:
            return None
    if p > 1:          # veio como 10 (=10%)
        p = p / 100.0
    return max(0.0, min(p, 0.95))


def _parse_estoque(v):
    """DISPONÍVEL / SIM / TRUE / 1 => True ; ESGOTADO / NÃO / 0 => False."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    s = _norm_nome(v)
    if any(k in s for k in ["DISPONIVEL", "SIM", "TRUE", "EM ESTOQUE", "ATIVO"]):
        return True
    if any(k in s for k in ["INDISPONIVEL", "ESGOTADO", "EM ESPERA", "NAO", "FALSE", "SEM ESTOQUE", "FALTA"]):
        return False
    return None


def carregar_excel(path: Path):
    """Lê a planilha com as colunas reais:
       PRODUTO | VOLUME | MEDIDA | ESTOQUE | Preço (R$) | PROMOÇÃO
       Retorna uma lista, preservando separadamente todas as variações de volume.
    """
    if not path.exists():
        raise FileNotFoundError(
            f"Planilha de estoque não encontrada em: {path}. "
            "A geração foi interrompida para não publicar preços fictícios."
        )
    try:
        import openpyxl
    except ImportError:
        print("\u26a0  openpyxl nao instalado (pip install openpyxl) - pulando Excel.")
        return {}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # localiza a linha de cabeçalho (pode não ser a linha 1)
    header_row, idx = None, {}
    for r in range(1, min(ws.max_row, 15) + 1):
        vals = [_norm_header(c.value) for c in ws[r]]
        mapa = {}
        for key, aliases in COL_ALIASES.items():
            for i, h in enumerate(vals):
                if h and (h in aliases or any(h.startswith(a) for a in aliases)):
                    mapa.setdefault(key, i)
                    break
        if "produto" in mapa and ("preco" in mapa or "estoque" in mapa):
            header_row, idx = r, mapa
            break

    if header_row is None:
        raise ValueError(
            "Cabeçalho da planilha não reconhecido. São necessárias as colunas "
            "PRODUTO, VOLUME, MEDIDA, ESTOQUE e Preço (R$)."
        )

    print(f"\u2713  Cabecalho na linha {header_row}: {idx}")

    def cell(row, key):
        i = idx.get(key)
        return row[i] if i is not None and i < len(row) else None

    dados = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        nome_raw = cell(row, "produto")
        if not nome_raw or not str(nome_raw).strip():
            continue
        nome_raw = str(nome_raw).strip()
        volume = cell(row, "volume")
        medida = cell(row, "medida")
        espec = None
        if volume not in (None, ""):
            v = str(volume).strip()
            if v.replace(".", "", 1).isdigit():
                v = str(int(float(v))) if float(v).is_integer() else v
            espec = f"{v} {str(medida).strip().lower()}" if medida else v
        # Quando PRODUTO já termina em "5 MG", remove somente essa repetição;
        # VOLUME + MEDIDA são a identificação oficial da variação.
        nome_base = re.sub(
            r"\s+\d+(?:[.,]\d+)?\s*(MG|ML|MCG|G|UI|IU)$", "", nome_raw,
            flags=re.IGNORECASE,
        ).strip()
        dados.append({
            "nome_planilha": nome_raw.upper(),
            "nome_base": nome_base.upper(),
            "espec": espec,
            "preco": _parse_preco(cell(row, "preco")),
            "promo_pct": _parse_promo(cell(row, "promocao")),
            "disponivel": _parse_estoque(cell(row, "estoque")),
        })
    print(f"\u2713  Excel lido: {len(dados)} produtos.")
    return dados


# =====================================================================
# 4) MAPEAMENTO DE IMAGENS (pasta imagens_produtos/)
# =====================================================================
IMG_EXTS = (".webp", ".png", ".jpg", ".jpeg", ".avif")


def _listar_imagens():
    if not IMG_DIR.exists():
        return []
    return [f for f in IMG_DIR.iterdir() if f.suffix.lower() in IMG_EXTS]


_IMGS = None


def encontrar_imagem(nome: str, espec: str) -> str:
    """Casa o produto com o arquivo da pasta imagens_produtos/ (ex.: 'AOD 9604 5 MG.webp')."""
    global _IMGS
    if _IMGS is None:
        _IMGS = _listar_imagens()
        print(f"\u2713  Imagens encontradas: {len(_IMGS)}")
    if not _IMGS:
        return ""
    tokens = [t for t in _norm_nome(nome).split() if t]
    num = "".join(c for c in str(espec) if c.isdigit())
    melhor, melhor_score = None, 0
    for f in _IMGS:
        fname = _norm_nome(f.stem)
        ftok = fname.split()
        score = sum(2 if t in ftok else (1 if t in fname else 0) for t in tokens)
        if score == 0:
            continue
        if num and num in ftok:
            score += 3
        score += 1 if len(ftok) == len(tokens) else 0
        if score > melhor_score:
            melhor, melhor_score = f, score
    return f"imagens_produtos/{melhor.name}" if melhor else ""


# =====================================================================
# 5) MONTAGEM DA LISTA FINAL DE PRODUTOS
# =====================================================================
def montar_produtos(dados_planilha: list):
    """Monta o catálogo exclusivamente pelas linhas da planilha.

    A base interna fornece apenas categoria, ícone e texto técnico. Preço,
    desconto, volume, medida e estoque vêm sempre da planilha selecionada.
    """
    if not dados_planilha:
        raise ValueError("A planilha não contém produtos; site não gerado.")

    metadados = []
    for nome, espec, _po, _pr, _pp, cat, icon, _disp, info in PRODUTOS_PADRAO:
        chave = _norm_nome(nome)
        # Nome histórico do catálogo; a identificação comercial da planilha é SLU PP.
        if chave == "SLU PP 332":
            chave = "SLU PP"
        metadados.append((chave, cat, icon, info))

    produtos = []
    for ov in dados_planilha:
        nome = ov["nome_base"]
        espec = ov.get("espec") or ""
        preco_base = ov.get("preco")
        if preco_base is None:
            print(f"⚠  Linha ignorada sem preço válido: {ov['nome_planilha']} {espec}")
            continue
        promo = ov.get("promo_pct")
        promo = promo if promo is not None else 0.0
        disponivel = ov.get("disponivel")
        disponivel = disponivel if disponivel is not None else False

        chave = _norm_nome(nome)
        meta = next((m for m in metadados if m[0] == chave), None)
        cat, icon, info = (meta[1], meta[2], meta[3]) if meta else (
            "Metabolismo", "\U0001f9ea",
            "Produto de pesquisa G-LAB. Consulte a ficha técnica no WhatsApp.",
        )
        # Emoji específico por função estudada (sobrepõe o ícone da categoria)
        icon = ICONES_PRODUTO.get(chave, icon)
        produtos.append({
            "nome": nome, "espec": espec,
            "precoOrig": preco_base,
            "preco": round(preco_base * (1 - promo), 2),
            "promoPct": promo, "cat": cat, "icon": icon, "info": info,
            "available": bool(disponivel),
            "img": encontrar_imagem(nome, espec),
        })

    for i, p in enumerate(produtos):
        p["id"] = i

    print("\n" + "=" * 84)
    print("CONFERÊNCIA — 100% dos valores abaixo vieram da planilha")
    print("=" * 84)
    print(f"{'PRODUTO':<34}{'ESPEC':<10}{'PREÇO':>12}{'PROMO':>9}{'FINAL':>12}")
    for p in produtos:
        print(f"{p['nome'][:33]:<34}{str(p['espec'])[:9]:<10}"
              f"{p['precoOrig']:>12.2f}{p['promoPct'] * 100:>8.2f}%"
              f"{p['preco']:>12.2f}")
    print("=" * 84 + "\n")

    print(f"\u2713  Catalogo montado: {len(produtos)} produtos "
          f"({sum(1 for p in produtos if p['available'])} disponiveis).")
    return produtos


# =====================================================================
# 6) CUPONS E REGIÕES 
# =====================================================================
CUPONS = {
    'BRUNA5': 0.05,'GILMARA5':0.05,'DAFNE10':0.10,'NOS5':0.05,'ROGERIO5':0.05,
      'ANDERSON5':0.05,'JAQUE5':0.05,'CABRAL5':0.05,'KARLINHA5':0.05,'LUD5':0.05,'CASSIA5':0.05,
      'THAIS5':0.05,'NATAN':0.00000000001,'LIRICY5':0.05,'ANDREAFLEURY':0.05,'ANA5':0.05,
      '10PRO':0.000000000001,'PRO5':0.05,'WEY5':0.05,'ALE5':0.05,'TRIGUEIRO':0.05,
      'RAYSSA5':0.05,'PATRICIA5':0.05,'LU5':0.05, 'RAFA5':0.05, 'WAWA':0.05, 'DUDA5':0.05, 
      'ALYNE5':0.05, 'JRCREMONEZ':0.05, 'ZAMA5':0.05, 'JENNI5':0.05, 'DJU5':0.05, 'CLAU5':0.05, 
      'GLAB5':0.05, 'BRENDA5':0.05,
}

REGIOES = {
    "SUL":          {"ufs": ["PR", "SC", "RS"],                                         "frete": 60.0,  "prazo": "3-9 dias"},
    "SUDESTE":      {"ufs": ["SP", "RJ", "MG", "ES"],                                   "frete": 90.0,  "prazo": "5-15 dias"},
    "CENTRO-OESTE": {"ufs": ["DF", "GO", "MT", "MS"],                                   "frete": 90.0,  "prazo": "5-15 dias"},
    "NORTE":        {"ufs": ["AM", "RR", "AP", "PA", "TO", "RO", "AC"],                 "frete": 110.0, "prazo": "10-30 dias"},
    "NORDESTE":     {"ufs": ["BA", "SE", "AL", "PE", "PB", "RN", "CE", "PI", "MA"],     "frete": 110.0, "prazo": "10-30 dias"},
}

# Curitiba + Regiao Metropolitana => frete regional R$ 20,00 (1-3 dias)
_RMC = [
    "CURITIBA", "SAO JOSE DOS PINHAIS", "COLOMBO", "PINHAIS", "ARAUCARIA",
    "ALMIRANTE TAMANDARE", "CAMPO LARGO", "CAMPO MAGRO", "FAZENDA RIO GRANDE",
    "PIRAQUARA", "QUATRO BARRAS", "CAMPINA GRANDE DO SUL", "ITAPERUCU",
    "RIO BRANCO DO SUL", "MANDIRITUBA", "CONTENDA", "BALSA NOVA", "LAPA",
    "TIJUCAS DO SUL", "AGUDOS DO SUL", "QUITANDINHA", "BOCAIUVA DO SUL",
    "ADRIANOPOLIS", "CERRO AZUL", "DOUTOR ULYSSES", "TUNAS DO PARANA",
    "PIEN", "RIO NEGRO", "CAMPO DO TENENTE",
]
FRETES_CIDADES = {f"{c}-PR": 20.0 for c in _RMC}

# =====================================================================
# 7) TEMPLATE HTML (SPA embutida)
# =====================================================================
HTML_TEMPLATE = r"""<!doctype html>
<html lang="pt-BR">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1, viewport-fit=cover"/>
<title>G-LAB PEPTIDES — Pesquisa & Longevidade</title>
<meta name="description" content="G-LAB Peptides — catálogo de peptídeos para pesquisa."/>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;600;700;800&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
<style>
:root{
  --red:#e10600; --red-2:#ff2d20; --red-glow:#ff3b30;
  --bg:#0a0a0a; --panel:#141414; --panel-2:#1c1c1c;
  --fg:#ffffff; --muted:#a0a0a0; --border:#2a2a2a;
  --font:'Space Grotesk',system-ui,sans-serif; --mono:'JetBrains Mono',monospace;
}
*{box-sizing:border-box;margin:0;padding:0}
html,body{background:var(--bg);color:var(--fg);font-family:var(--font);-webkit-font-smoothing:antialiased;overflow-x:hidden}
body{background:
  radial-gradient(ellipse 60% 40% at 20% 0%,rgba(225,6,0,.25),transparent 60%),
  radial-gradient(ellipse 50% 40% at 90% 20%,rgba(255,45,32,.18),transparent 60%),
  var(--bg);background-attachment:fixed}
a{color:inherit;text-decoration:none}
button{font-family:inherit;cursor:pointer;border:none;background:none;color:inherit}
img{max-width:100%;display:block}

/* ── HEADER ── */
.hdr{position:sticky;top:0;z-index:50;backdrop-filter:blur(14px);background:rgba(10,10,10,.75);border-bottom:1px solid var(--border)}
.hdr-inner{max-width:1400px;margin:0 auto;padding:14px 20px;display:flex;align-items:center;gap:16px;justify-content:space-between}
.logo{display:flex;align-items:center;gap:10px;font-weight:800;font-size:1.25rem;letter-spacing:.02em}
.logo-hex{width:36px;height:40px;display:grid;place-items:center;color:var(--red);
  clip-path:polygon(50% 0,100% 25%,100% 75%,50% 100%,0 75%,0 25%);
  background:linear-gradient(135deg,var(--red),var(--red-2));font-weight:900;font-size:1.4rem;color:#fff}
.logo strong{color:#fff} .logo em{color:var(--red);font-style:normal}
.hdr-cart{display:flex;align-items:center;gap:8px;padding:10px 16px;border-radius:999px;background:var(--red);color:#fff;font-weight:700;font-size:.9rem}
.hdr-cart:hover{background:var(--red-2)}
.cart-count{background:#fff;color:var(--red);border-radius:999px;padding:1px 8px;font-size:.75rem;font-weight:800;min-width:20px;text-align:center}

/* ── HERO / ORBIT ── */
.hero{position:relative;padding:24px 16px 40px;text-align:center}
.hero h1{font-size:clamp(1.8rem,5vw,3.2rem);font-weight:800;letter-spacing:-.02em;line-height:1.05}
.hero h1 span{background:linear-gradient(120deg,var(--red),#fff);-webkit-background-clip:text;background-clip:text;color:transparent}
.hero p{color:var(--muted);margin-top:8px;font-size:.95rem}
.orbit-stage{position:relative;width:100%;height:420px;perspective:1400px;perspective-origin:50% 40%;margin-top:8px;--orbit-scale:1}
@media(min-width:768px){.orbit-stage{height:540px}}
.orbit-ring{position:absolute;inset:0;transform-style:preserve-3d;transition:transform .9s cubic-bezier(.22,1,.36,1);animation:spin 30s linear infinite;
  scale:var(--orbit-scale)}
.orbit-ring.paused{animation-play-state:paused}

@keyframes spin{from{transform:rotateY(0)}to{transform:rotateY(-360deg)}}
.orbit-item{position:absolute;top:50%;left:50%;width:130px;height:220px;margin:-110px 0 0 -65px;transform-style:preserve-3d;cursor:pointer}
@media(min-width:768px){.orbit-item{width:160px;height:270px;margin:-135px 0 0 -80px}}
.orbit-can{position:relative;width:100%;height:100%;border-radius:18px;overflow:hidden;
  background:linear-gradient(135deg,#1a1a1a,#0a0a0a);
  border:1px solid color-mix(in oklab,var(--can-color) 60%,transparent);
  box-shadow:0 20px 40px -20px color-mix(in oklab,var(--can-color) 60%,transparent),0 0 30px color-mix(in oklab,var(--can-color) 25%,transparent);
  display:flex;flex-direction:column;align-items:center;justify-content:space-between;padding:12px 8px}
.orbit-can img{width:100%;height:70%;object-fit:contain;filter:drop-shadow(0 6px 12px rgba(0,0,0,.6))}
.orbit-can .fallback{font-size:2.4rem}
.orbit-can .tag{font-family:var(--mono);font-size:.55rem;letter-spacing:.16em;text-transform:uppercase;padding:3px 8px;border-radius:999px;background:rgba(0,0,0,.5);border:1px solid rgba(255,255,255,.12);color:#fff}
.orbit-can .name{font-weight:800;font-size:.75rem;color:#fff;text-align:center;line-height:1.1;padding:0 4px}
.orbit-nav{position:absolute;top:50%;transform:translateY(-50%);width:44px;height:44px;border-radius:50%;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.16);backdrop-filter:blur(10px);color:#fff;display:grid;place-items:center;z-index:20;font-size:1.4rem}
.orbit-nav.prev{left:8px} .orbit-nav.next{right:8px}

/* ── FILTROS ── */
.filters{max-width:1400px;margin:0 auto;padding:16px 16px 0;display:flex;flex-wrap:wrap;gap:8px;align-items:center}
.search{flex:1;min-width:220px;padding:12px 16px;background:var(--panel);border:1px solid var(--border);border-radius:12px;color:#fff;font-size:.9rem}
.search::placeholder{color:var(--muted)}
.cat-pill{padding:8px 14px;border-radius:999px;border:1px solid rgba(255,255,255,.12);background:rgba(255,255,255,.03);color:rgba(255,255,255,.75);font-family:var(--mono);font-size:.7rem;letter-spacing:.12em;text-transform:uppercase;white-space:nowrap;transition:.2s}
.cat-pill:hover{color:#fff;border-color:rgba(255,255,255,.3)}
.cat-pill.active{color:#fff;background:var(--red);border-color:var(--red);box-shadow:0 0 20px rgba(225,6,0,.5)}

/* ── GRID DE PRODUTOS ── */
.grid{max-width:1400px;margin:0 auto;padding:20px 16px 60px;display:grid;grid-template-columns:repeat(auto-fill,minmax(220px,1fr));gap:14px}
.pcard{position:relative;border-radius:16px;padding:16px;background:linear-gradient(180deg,var(--panel),var(--panel-2));border:1px solid var(--border);transition:.3s;overflow:hidden;display:flex;flex-direction:column;gap:10px}
.pcard:hover{transform:translateY(-4px);border-color:var(--red);box-shadow:0 20px 40px -20px rgba(225,6,0,.4)}
.pcard.unavailable{opacity:.5}
.pcard-img{aspect-ratio:1;background:#000;border-radius:12px;display:grid;place-items:center;overflow:hidden;border:1px solid var(--border)}
.pcard-img img{width:100%;height:100%;object-fit:contain}
.pcard-img .fallback{font-size:3rem}
.pcard-tag{position:absolute;top:10px;right:10px;font-family:var(--mono);font-size:.55rem;letter-spacing:.14em;text-transform:uppercase;padding:4px 8px;border-radius:999px;background:rgba(0,0,0,.7);color:#fff;border:1px solid var(--red)}
.pcard-name{font-weight:800;font-size:1rem;line-height:1.2}
.pcard-emoji{font-size:1.05rem;margin-right:2px}
.pcard-spec{font-family:var(--mono);font-size:.7rem;color:var(--muted);letter-spacing:.1em}
.pcard-price{display:flex;align-items:baseline;gap:8px;margin-top:auto}
.pcard-price .cur{font-weight:800;font-size:1.2rem;color:var(--red)}
.pcard-price .old{color:var(--muted);text-decoration:line-through;font-size:.8rem}
.pcard-btn{padding:10px;background:var(--red);color:#fff;border-radius:10px;font-weight:700;font-size:.85rem;transition:.2s}
.pcard-btn:hover{background:var(--red-2)}
.cep-status{font-family:var(--mono);font-size:.7rem;color:var(--muted);margin:6px 0 2px;min-height:14px}
.toast{position:fixed;left:50%;bottom:24px;transform:translateX(-50%) translateY(20px);z-index:200;background:var(--red);color:#fff;font-weight:700;font-size:.85rem;padding:12px 20px;border-radius:999px;box-shadow:0 12px 30px -10px rgba(225,6,0,.7);opacity:0;pointer-events:none;transition:.25s}
.toast.show{opacity:1;transform:translateX(-50%) translateY(0)}
.hdr-cart.pulse{animation:cartPulse .45s ease}
/* FAB carrinho flutuante */
.cart-fab{position:fixed;right:18px;bottom:18px;z-index:90;display:flex;align-items:center;gap:8px;
  padding:14px 18px;border-radius:999px;background:var(--red);color:#fff;font-weight:800;font-size:.9rem;
  border:2px solid rgba(255,255,255,.15);box-shadow:0 10px 30px rgba(225,6,0,.45),0 4px 14px rgba(0,0,0,.6);
  cursor:pointer;transition:transform .18s ease,box-shadow .18s ease,opacity .2s ease}
.cart-fab:hover{transform:translateY(-3px) scale(1.04);background:var(--red-2)}
.cart-fab .fab-ico{font-size:1.25rem;line-height:1}
.cart-fab .fab-count{position:absolute;top:-6px;right:-6px;background:#fff;color:var(--red);border-radius:999px;
  min-width:24px;height:24px;display:grid;place-items:center;font-size:.75rem;font-weight:900;
  box-shadow:0 2px 8px rgba(0,0,0,.5);padding:0 6px}
.cart-fab.empty{opacity:.55}
.cart-fab.pulse{animation:cartPulse .45s ease}
/* WhatsApp flutuante */
.wa-fab{position:fixed;right:18px;bottom:84px;z-index:90;display:flex;align-items:center;gap:8px;
  padding:13px 17px;border-radius:999px;background:#25d366;color:#04220f;font-weight:800;font-size:.88rem;
  border:2px solid rgba(255,255,255,.18);box-shadow:0 10px 30px rgba(37,211,102,.35),0 4px 14px rgba(0,0,0,.6);
  cursor:pointer;transition:transform .18s ease,background .18s ease;text-decoration:none}
.wa-fab:hover{transform:translateY(-3px) scale(1.04);background:#1eb356}
.wa-fab svg{width:22px;height:22px;flex:none;fill:currentColor}
@media(max-width:640px){
  .cart-fab{right:14px;bottom:14px;padding:13px 15px;font-size:.82rem}
  .cart-fab .fab-label{display:none}
  .wa-fab{right:14px;bottom:78px;padding:12px 14px;font-size:.8rem}
  .wa-fab .fab-label{display:none}
}
@media print{.cart-fab,.wa-fab{display:none}}
@keyframes cartPulse{0%{transform:scale(1)}40%{transform:scale(1.14)}100%{transform:scale(1)}}
.pcard-btn:disabled{background:#333;cursor:not-allowed;opacity:.6}

/* ── MODAL BASE ── */
.modal{position:fixed;inset:0;z-index:100;background:rgba(0,0,0,.85);backdrop-filter:blur(8px);display:none;align-items:center;justify-content:center;padding:16px;overflow-y:auto}
.modal.open{display:flex}
.modal-panel{background:var(--panel);border:1px solid var(--border);border-radius:20px;max-width:520px;width:100%;padding:24px;position:relative;max-height:90vh;overflow-y:auto;-webkit-overflow-scrolling:touch}
.modal-panel.large{max-width:720px}
.modal-close{position:absolute;top:12px;right:12px;width:36px;height:36px;border-radius:50%;background:rgba(255,255,255,.08);color:#fff;font-size:1.2rem;display:grid;place-items:center;z-index:2}
.modal-close:hover{background:var(--red)}
.modal h2{font-size:1.4rem;font-weight:800;margin-bottom:8px;padding-right:40px}
.modal .subtle{color:var(--muted);font-size:.9rem;margin-bottom:16px}

/* ── WELCOME (mobile-friendly, com scroll interno) ── */
#welcome .modal-panel{max-height:88vh;display:flex;flex-direction:column;text-align:center}
#welcome .welcome-scroll{overflow-y:auto;flex:1;padding-right:4px;margin-bottom:16px}
#welcome .welcome-eyebrow{color:var(--red);font-size:.72rem;font-weight:800;letter-spacing:.28em;text-transform:uppercase;margin-bottom:8px}
#welcome h2{font-size:1.7rem;font-weight:800;padding-right:0;margin-bottom:12px}
#welcome .welcome-rule{width:56px;height:3px;background:var(--red);border-radius:3px;margin:0 auto 18px}
#welcome .welcome-list{display:grid;gap:12px;margin-top:4px;text-align:left}
#welcome .welcome-list li{display:flex;gap:14px;align-items:center;padding:14px;border-radius:14px;background:rgba(255,255,255,.04);border:1px solid var(--border);font-size:.85rem;line-height:1.4}
#welcome .welcome-list .wi{width:40px;height:40px;flex-shrink:0;border-radius:10px;background:rgba(255,255,255,.06);display:grid;place-items:center;font-size:1.15rem}
#welcome .welcome-list .wt{font-weight:800;color:#fff;margin-bottom:2px;font-size:.95rem}
#welcome .welcome-list .wd{color:var(--muted);font-size:.85rem}
#welcome .welcome-list b{color:#fff}
#welcome .btn-enter{width:100%;padding:16px;background:var(--red);color:#fff;font-weight:800;border-radius:12px;font-size:1rem;letter-spacing:.14em;text-transform:uppercase;flex-shrink:0}
#welcome .btn-enter:hover{background:var(--red-2)}
/* ── GUIAS (PDF) ── */
.guides{display:flex;flex-wrap:wrap;gap:12px;justify-content:center;padding:24px 16px 4px}
.guide-btn{display:inline-flex;align-items:center;gap:10px;padding:13px 20px;border-radius:12px;border:1px solid var(--border);background:rgba(255,255,255,.04);color:#fff;font-weight:800;font-size:.85rem;letter-spacing:.06em;text-transform:uppercase;text-decoration:none;transition:.2s}
.guide-btn:hover{background:var(--red);border-color:var(--red)}


/* ── CARRINHO ── */
#cart .modal-panel{max-width:520px}
.cart-item{display:flex;gap:10px;padding:10px 0;border-bottom:1px solid var(--border);align-items:center}
.cart-item img{width:44px;height:44px;border-radius:8px;object-fit:contain;background:#000}
.cart-item .info{flex:1;min-width:0}
.cart-item .name{font-weight:700;font-size:.85rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.cart-item .price{color:var(--red);font-weight:700;font-size:.8rem}
.cart-item.brinde{background:linear-gradient(90deg,rgba(225,6,0,.1),transparent);border-radius:8px;padding-left:8px}
.cart-item.brinde .tag-brinde{background:var(--red);color:#fff;font-size:.6rem;padding:2px 6px;border-radius:4px;font-weight:800;letter-spacing:.08em;margin-left:4px}
.qty{display:flex;align-items:center;gap:6px}
.qty button{width:26px;height:26px;border-radius:6px;background:var(--panel-2);color:#fff;font-weight:700;border:1px solid var(--border)}
.qty span{min-width:20px;text-align:center;font-weight:700;font-size:.85rem}
.rm{color:var(--red);font-size:.75rem;margin-left:8px}
.cart-inputs{display:grid;gap:8px;margin:12px 0}
.cart-inputs input{padding:10px 12px;background:var(--panel-2);border:1px solid var(--border);border-radius:8px;color:#fff;font-size:.85rem;font-family:var(--mono)}
.cart-totals{display:grid;gap:6px;padding:12px;border-radius:10px;background:rgba(255,255,255,.03);border:1px solid var(--border);margin-bottom:12px}
.cart-totals .row{display:flex;justify-content:space-between;font-size:.85rem}
.cart-totals .row.tot{font-size:1.1rem;font-weight:800;color:var(--red);border-top:1px solid var(--border);padding-top:8px;margin-top:4px}
.btn-checkout{width:100%;padding:14px;background:var(--red);color:#fff;font-weight:800;border-radius:10px;font-size:.95rem;letter-spacing:.04em;text-transform:uppercase}
.btn-checkout:hover{background:var(--red-2)}
.btn-checkout:disabled{background:#333;cursor:not-allowed}
.brinde-warn{padding:10px;background:rgba(225,6,0,.12);border:1px solid var(--red);border-radius:8px;font-size:.8rem;color:#fff;margin:8px 0;line-height:1.4}
.brinde-warn b{color:var(--red)}

/* ── MODAL DE FRETE OBRIGATÓRIO ── */
#shipReq .modal-panel{max-width:420px;text-align:center}
#shipReq .icon{font-size:3rem;margin-bottom:8px}
#shipReq h2{color:var(--red);text-align:center;padding-right:0}
#shipReq p{color:var(--muted);margin:8px 0 16px;font-size:.9rem;line-height:1.4}
#shipReq input{width:100%;padding:12px;background:var(--panel-2);border:1px solid var(--border);border-radius:10px;color:#fff;font-family:var(--mono);text-align:center;font-size:1.1rem;letter-spacing:.08em;margin-bottom:12px}
#shipReq .btns{display:flex;gap:8px}
#shipReq .btns button{flex:1;padding:12px;border-radius:10px;font-weight:700}
#shipReq .btn-cep{background:var(--red);color:#fff}
#shipReq .btn-back{background:var(--panel-2);color:#fff;border:1px solid var(--border)}

/* ── BOTÕES DO MODAL DE IDENTIFICAÇÃO ── */
#ident .btns{display:flex;flex-direction:column;gap:18px;margin-top:22px}
#ident .btns button{width:100%;padding:16px;border-radius:12px;font-weight:800;font-size:1rem;letter-spacing:.1em;text-transform:uppercase;cursor:pointer}
#ident .btn-send{background:var(--red);color:#fff;border:1px solid var(--red);box-shadow:0 10px 26px rgba(225,6,0,.35);order:-1}
#ident .btn-send:hover{background:var(--red-2)}
#ident .btn-back{background:transparent;color:var(--muted);border:1px solid var(--border);font-size:.82rem;letter-spacing:.06em}
#ident .btn-back:hover{color:#fff;background:rgba(255,255,255,.05)}

/* ── CERTIFICADOS ── */
#certs .modal-panel{max-width:460px}
#certs .cert-list{display:grid;gap:10px;margin-top:14px}
#certs .cert-list a{display:flex;align-items:center;gap:10px;padding:14px;border-radius:12px;border:1px solid var(--border);background:rgba(255,255,255,.04);color:#fff;font-weight:700;text-decoration:none;font-size:.9rem}
#certs .cert-list a:hover{background:var(--red);border-color:var(--red)}

/* ── DETALHE (dossiê) ── */
#detail .modal-panel{max-width:800px}
#detail .detail-hero{display:grid;grid-template-columns:1fr;gap:16px;margin-bottom:16px}
@media(min-width:640px){#detail .detail-hero{grid-template-columns:200px 1fr}}
#detail .detail-img{aspect-ratio:1;background:#000;border-radius:14px;display:grid;place-items:center;overflow:hidden;border:1px solid var(--border)}
#detail .detail-img img{width:100%;height:100%;object-fit:contain}
#detail .detail-img .fallback{font-size:4rem}
#detail .detail-info h2{color:var(--red)}
#detail .detail-price{display:flex;align-items:baseline;gap:10px;margin:10px 0}
#detail .detail-price .cur{font-size:2rem;font-weight:800;color:var(--red)}
#detail .detail-price .old{text-decoration:line-through;color:var(--muted)}
#detail .body-txt{color:#ddd;line-height:1.6;font-size:.9rem;padding:12px;background:rgba(255,255,255,.03);border-radius:10px;border:1px solid var(--border)}

/* ── FOOTER / TICKER ── */
.ticker{overflow:hidden;border-top:1px solid var(--border);border-bottom:1px solid var(--border);padding:10px 0;background:#000;white-space:nowrap;font-family:var(--mono);font-size:.75rem;letter-spacing:.14em;color:var(--muted)}
.ticker-inner{display:inline-block;animation:tk 40s linear infinite;padding-left:100%}
.ticker span{margin:0 20px}
.ticker .red{color:var(--red);font-weight:700}
@keyframes tk{from{transform:translateX(0)}to{transform:translateX(-100%)}}
.foot{padding:24px 16px;text-align:center;color:var(--muted);font-size:.75rem}

/* ── RESPONSIVO MOBILE ── */
@media(max-width:640px){
  .hdr-inner{padding:12px 14px}
  .logo{font-size:1.05rem}
  .hdr-cart{padding:8px 12px;font-size:.8rem}
  .grid{grid-template-columns:repeat(2,1fr);gap:10px;padding:16px 12px 50px}
  .pcard{padding:12px}
  .pcard-name{font-size:.85rem}
  .pcard-price .cur{font-size:1rem}
  .modal-panel{padding:18px;border-radius:16px;max-height:92vh}
  .modal h2{font-size:1.15rem}
  #welcome .welcome-list li{font-size:.78rem;padding:8px}
  .cart-item .name{font-size:.78rem}
  .filters{padding:12px 12px 0}
  .search{font-size:.85rem;padding:10px 14px}
}

/* ══════════ CALCULADORA DE PEPTÍDEOS ══════════ */
.calc-wrap{max-width:1100px;margin:26px auto 6px;padding:0 14px}
.calc-card{background:linear-gradient(180deg,rgba(255,255,255,.045),rgba(255,255,255,.015));
  border:1px solid rgba(255,255,255,.10);border-radius:20px;padding:20px 16px;position:relative;overflow:hidden}
.calc-card::before{content:"";position:absolute;inset:0;pointer-events:none;
  background:radial-gradient(ellipse 90% 55% at 50% 0%,rgba(225,6,0,.20),transparent 62%)}
.calc-card h2{position:relative;font-size:1.35rem;font-weight:900;letter-spacing:.02em;text-align:center;margin:0 0 4px}
.calc-card h2 span{color:var(--red)}
.calc-sub{position:relative;text-align:center;font-family:var(--mono);font-size:.68rem;letter-spacing:.18em;
  text-transform:uppercase;color:rgba(255,255,255,.55);margin-bottom:16px}
.calc-grid{position:relative;display:grid;gap:16px}
@media(min-width:900px){.calc-grid{grid-template-columns:1.05fr .95fr;align-items:start}}
.calc-step{margin-bottom:14px}
.calc-step-t{font-family:var(--mono);font-size:.66rem;letter-spacing:.16em;text-transform:uppercase;
  color:rgba(255,255,255,.6);margin-bottom:8px;display:flex;align-items:center;gap:8px}
.calc-step-t b{display:grid;place-items:center;width:20px;height:20px;border-radius:50%;
  background:var(--red);color:#fff;font-size:.62rem}
.calc-opts{display:flex;flex-wrap:wrap;gap:8px}
.calc-opt{padding:9px 13px;border-radius:999px;cursor:pointer;font-family:var(--mono);font-size:.72rem;
  letter-spacing:.06em;color:rgba(255,255,255,.8);background:rgba(255,255,255,.04);
  border:1px solid rgba(255,255,255,.12);transition:all .2s;white-space:nowrap}
.calc-opt:hover{color:#fff;border-color:rgba(255,255,255,.35);transform:translateY(-1px)}
.calc-opt.active{background:var(--red);border-color:var(--red);color:#fff;font-weight:700;
  box-shadow:0 0 18px rgba(225,6,0,.45)}
.calc-custom{display:flex;align-items:center;gap:6px;background:rgba(255,255,255,.04);
  border:1px solid rgba(255,255,255,.12);border-radius:999px;padding:4px 10px}
.calc-custom input{width:70px;background:transparent;border:0;color:#fff;font-family:var(--mono);
  font-size:.75rem;outline:none;text-align:center}
.calc-custom span{font-family:var(--mono);font-size:.62rem;color:rgba(255,255,255,.5)}
.calc-res{background:rgba(0,0,0,.35);border:1px solid rgba(255,255,255,.10);border-radius:16px;padding:16px}
.calc-big{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:14px}
.calc-big div{flex:1 1 120px;background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.10);
  border-radius:12px;padding:10px 12px}
.calc-big small{display:block;font-family:var(--mono);font-size:.58rem;letter-spacing:.14em;
  text-transform:uppercase;color:rgba(255,255,255,.5);margin-bottom:4px}
.calc-big b{font-size:1.3rem;font-weight:900;color:#fff}
.calc-big b em{font-style:normal;font-size:.75rem;color:rgba(255,255,255,.6);font-weight:600}
.calc-big.hl b{color:var(--red)}
.syr{position:relative;margin:6px 0 4px}
.syr svg{width:100%;height:auto;display:block;overflow:visible}
.calc-note{font-size:.74rem;line-height:1.5;color:rgba(255,255,255,.6);margin-top:10px}
.calc-warn{margin-top:10px;font-size:.76rem;font-weight:700;color:#ffb4b0;background:rgba(225,6,0,.12);
  border:1px solid rgba(225,6,0,.35);border-radius:10px;padding:9px 11px}
.calc-actions{display:flex;flex-wrap:wrap;gap:8px;margin-top:12px}
</style>
</head>
<body>

<!-- HEADER -->
<header class="hdr">
  <div class="hdr-inner">
    <div class="logo">
      <div class="logo-hex">G</div>
      <span><strong>G-LAB</strong> <em>PEPTIDES</em></span>
    </div>
    <button class="hdr-cart" id="cartBtn" data-action="open-cart">
      🛒 CARRINHO <span class="cart-count" id="cartCount">0</span>
    </button>
  </div>
</header>

<!-- WHATSAPP FLUTUANTE (duvidas) -->
<a class="wa-fab" id="waFab" href="https://wa.me/17746222523?text=Ol%C3%A1%21%20Tenho%20uma%20d%C3%BAvida%20sobre%20os%20pept%C3%ADdeos%20da%20G-LAB."
   target="_blank" rel="noopener" aria-label="Falar no WhatsApp" title="Tire suas dúvidas no WhatsApp"
   onclick="try{track('clicou_whatsapp_duvida')}catch(e){}">
  <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M12.04 2C6.58 2 2.13 6.45 2.13 11.91c0 1.75.46 3.45 1.32 4.95L2 22l5.25-1.38a9.9 9.9 0 0 0 4.79 1.22h.01c5.46 0 9.91-4.45 9.91-9.91C21.96 6.45 17.5 2 12.04 2zm5.8 14.03c-.24.68-1.4 1.3-1.94 1.35-.5.05-1.13.07-1.82-.11-.42-.11-.96-.29-1.65-.59-2.9-1.25-4.79-4.17-4.94-4.37-.14-.19-1.18-1.57-1.18-3 0-1.43.75-2.13 1.02-2.42.27-.29.58-.36.78-.36h.56c.18 0 .42-.07.66.5.24.58.82 2.01.89 2.15.07.14.12.31.02.5-.09.19-.14.31-.28.48-.14.17-.29.37-.42.5-.14.14-.28.29-.12.57.16.29.71 1.17 1.52 1.9 1.05.93 1.93 1.22 2.2 1.36.27.14.43.12.59-.07.16-.19.68-.79.86-1.06.18-.27.36-.22.6-.13.24.09 1.53.72 1.79.85.26.14.43.21.5.33.07.12.07.68-.17 1.36z"/></svg>
  <span class="fab-label">DÚVIDAS</span>
</a>

<!-- CARRINHO FLUTUANTE (scroll em toda a pagina) -->
<button class="cart-fab empty" id="cartFab" data-action="open-cart" aria-label="Abrir carrinho" title="Ver meu pedido">
  <span class="fab-ico">&#128722;</span><span class="fab-label">MEU PEDIDO</span>
  <span class="fab-count" id="cartFabCount">0</span>
</button>

<!-- HERO -->
<section class="hero">
  <h1>PEPTÍDEOS DE <span>ALTA PUREZA</span></h1>
  <p>Catálogo científico • Uso em pesquisa • Envio nacional</p>
  <div class="orbit-stage" id="orbitStage">
    <button class="orbit-nav prev" onclick="rotOrbit(-1)">‹</button>
    <div class="orbit-ring" id="orbitRing"></div>
    <button class="orbit-nav next" onclick="rotOrbit(1)">›</button>
  </div>
</section>

<!-- CALCULADORA DE PEPTÍDEOS -->
<section class="calc-wrap" id="calculadora">
  <div class="calc-card">
    <h2>CALCULADORA DE <span>PEPTÍDEOS</span></h2>
    <div class="calc-sub">Dose • Frasco • Diluente • Seringa</div>
    <div class="calc-grid">
      <div>
        <div class="calc-step">
          <div class="calc-step-t"><b>1</b> Dose desejada</div>
          <div class="calc-opts" id="calcDoseOpts"></div>
          <div class="calc-opts" style="margin-top:8px">
            <label class="calc-custom">
              <input id="calcDoseCustom" type="number" min="0" step="any" placeholder="outra"/>
              <span id="calcDoseUnit">mcg</span>
            </label>
            <button type="button" class="calc-opt" id="calcUnitMcg" onclick="calcSetUnit('mcg')">mcg</button>
            <button type="button" class="calc-opt" id="calcUnitMg" onclick="calcSetUnit('mg')">mg</button>
          </div>
        </div>
        <div class="calc-step">
          <div class="calc-step-t"><b>2</b> Volume / conteúdo do frasco (mg)</div>
          <div class="calc-opts" id="calcVialOpts"></div>
          <div class="calc-opts" style="margin-top:8px">
            <label class="calc-custom">
              <input id="calcVialCustom" type="number" min="0" step="any" placeholder="outro"/>
              <span>mg</span>
            </label>
          </div>
        </div>
        <div class="calc-step">
          <div class="calc-step-t"><b>3</b> Diluente adicionado (ml de água bacteriostática)</div>
          <div class="calc-opts" id="calcDilOpts"></div>
          <div class="calc-opts" style="margin-top:8px">
            <label class="calc-custom">
              <input id="calcDilCustom" type="number" min="0" step="any" placeholder="outro"/>
              <span>ml</span>
            </label>
          </div>
        </div>
        <div class="calc-step" style="margin-bottom:0">
          <div class="calc-step-t"><b>4</b> Seringa de insulina</div>
          <div class="calc-opts" id="calcSyrOpts"></div>
        </div>
      </div>
      <div class="calc-res">
        <div class="calc-big">
          <div class="hl"><small>Aplicar</small><b id="calcUI">—</b></div>
          <div><small>Volume</small><b id="calcML">—</b></div>
          <div><small>Concentração</small><b id="calcConc">—</b></div>
        </div>
        <div class="syr" id="calcSyr"></div>
        <div id="calcExtra" class="calc-note"></div>
        <div id="calcWarn"></div>
        <div class="calc-actions">
          <button type="button" class="calc-opt" onclick="calcReset()">↺ Limpar</button>
          <a class="calc-opt" href="reconstituicao/1.pdf" target="_blank" rel="noopener">💧 Guia de reconstituição</a>
          <a class="calc-opt" id="calcZap" href="#" target="_blank" rel="noopener">💬 Tirar dúvida</a>
        </div>
        <div class="calc-note">Cálculo de referência para uso em pesquisa. <b>1 ml = 100 UI</b> na seringa U-100. Confira sempre o rótulo do frasco.</div>
      </div>
    </div>
  </div>
</section>

<!-- TICKER -->
<div class="ticker"><div class="ticker-inner">
  <span class="red">● G-LAB PEPTIDES</span><span>PUREZA >99%</span>
  <span class="red">● ENVIO NACIONAL</span><span>PAGAMENTO PIX</span>
  <span class="red">● FRETE GRÁTIS ACIMA DE R$2.000</span>
  <span>BRINDE ACIMA DE R$1.000: BACTERIOSTATIC WATER POR ITEM</span>
  <span class="red">● SUPORTE WHATSAPP</span>
</div></div>

<!-- GUIAS EM PDF -->
<div class="guides">
  <a class="guide-btn" href="conversao/1.pdf" target="_blank" rel="noopener"
     onclick="try{track('abriu_guia_calculo')}catch(e){}">🧮 Guia de cálculo / conversão</a>
  <a class="guide-btn" href="reconstituicao/1.pdf" target="_blank" rel="noopener"
     onclick="try{track('abriu_guia_reconstituicao')}catch(e){}">💧 Como fazer a reconstituição</a>
  <button class="guide-btn" type="button" onclick="openModal('certs')">🔬 Certificados de pureza</button>
</div>

<!-- FILTROS -->
<div class="filters" id="filters"></div>

<!-- GRID -->
<main class="grid" id="grid"></main>

<!-- FOOTER -->
<footer class="foot">
  <div class="guides" style="padding-top:0;padding-bottom:18px">
    <a class="guide-btn" href="conversao/1.pdf" target="_blank" rel="noopener">🧮 Guia de cálculo</a>
    <a class="guide-btn" href="reconstituicao/1.pdf" target="_blank" rel="noopener">💧 Guia de reconstituição</a>
    <button class="guide-btn" type="button" onclick="openModal('certs')">🔬 Certificados de pureza</button>
  </div>
  © G-LAB PEPTIDES • Produtos destinados exclusivamente a pesquisa científica.<br>
  Não indicados para uso humano ou veterinário.
</footer>

<!-- MODAL WELCOME -->
<div class="modal open" id="welcome">
  <div class="modal-panel">
    <button class="modal-close" onclick="closeModal('welcome')">×</button>
    <div class="welcome-scroll">
      <p class="welcome-eyebrow">Bem-vindo à G-LAB</p>
      <h2>Por que escolher a <span style="color:var(--red)">G-LAB</span>?</h2>
      <div class="welcome-rule"></div>
      <ul class="welcome-list">
        <li><span class="wi">🤝</span><span><span class="wt">Atendimento Personalizado: </span><span class="wd">Suporte dedicado e exclusivo para cada cliente.</span></span></li>
        <li><span class="wi">🧬</span><span><span class="wt">Auxílio nos Protocolos: </span><span class="wd">Orientação especializada na montagem do seu protocolo.</span></span></li>
        <li><span class="wi">💬</span><span><span class="wt">Acompanhamento Completo: </span><span class="wd">Esclarecimento de dúvidas durante todo o tratamento.</span></span></li>
        <li><span class="wi">🛡️</span><span><span class="wt">Compra Garantida: </span><span class="wd">Se seu pacote não chegar ou for extraviado, o reembolso é <b>TOTAL</b>!</span></span></li>
      </ul>
    </div>
    <button class="btn-enter" onclick="closeModal('welcome')">Entrar no site ✨</button>
  </div>
</div>


<!-- MODAL DETAIL -->
<div class="modal" id="detail">
  <div class="modal-panel large">
    <button class="modal-close" onclick="closeModal('detail')">×</button>
    <div id="detailBody"></div>
  </div>
</div>

<!-- MODAL CART -->
<div class="modal" id="cart">
  <div class="modal-panel">
    <button class="modal-close" onclick="closeModal('cart')">×</button>
    <h2>Seu carrinho</h2>
    <div id="cartItems"></div>
    <div id="brindeMsg"></div>
    <div class="cart-inputs">
      <input id="cepInput" placeholder="CEP para cálculo de frete (só números)" maxlength="9" oninput="onCepInput()" inputmode="numeric"/>
      <input id="cupomInput" placeholder="Cupom de desconto" oninput="calcTotals()"/>
    </div>
    <div class="cep-status" id="cepStatus"></div>
    <div class="cart-totals" id="totals"></div>
    <button class="btn-checkout" id="btnCheckout" onclick="checkout()">Finalizar via WhatsApp</button>
  </div>
</div>

<!-- MODAL IDENTIFICACAO (nome + whatsapp antes de finalizar) -->
<div class="modal" id="ident">
  <div class="modal-panel">
    <button class="modal-close" onclick="closeIdent()">&times;</button>
    <div class="icon">&#128100;</div>
    <h2>Dados para entrega</h2>
    <p>Preencha seus dados para registrarmos o pedido e emitirmos a etiqueta de envio.</p>
    <input id="f_nome"   placeholder="Nome completo *" maxlength="120" autocomplete="name"/>
    <input id="f_cpf"    placeholder="CPF *" maxlength="14" inputmode="numeric" oninput="mascaraCPF(this)"/>
    <input id="f_tel"    placeholder="WhatsApp (DDD + n&uacute;mero) *" maxlength="20" inputmode="numeric" autocomplete="tel"/>
    <input id="f_end"    placeholder="Endere&ccedil;o (rua/avenida) *" maxlength="200" autocomplete="address-line1"/>
    <input id="f_num"    placeholder="N&uacute;mero *" maxlength="10"/>
    <input id="f_comp"   placeholder="Complemento" maxlength="100"/>
    <input id="f_bairro" placeholder="Bairro *" maxlength="100"/>
    <input id="f_cidade" placeholder="Cidade *" maxlength="100"/>
    <input id="f_estado" placeholder="UF *" maxlength="2"/>
    <input id="cep-destino" placeholder="CEP *" maxlength="9" inputmode="numeric"/>
    <select id="f_pgto">
      <option value="Pix">PIX</option>
      <option value="Cartao">CART&Atilde;O DE CR&Eacute;DITO</option>
    </select>
    <div class="ident-err" id="identErr"></div>
    <div class="btns">
      <button class="btn-send" type="button" onclick="confirmIdent()">✅ Enviar pedido</button>
      <button class="btn-back" type="button" onclick="closeIdent()">&#10005; Voltar ao carrinho</button>
    </div>
  </div>
</div>

<!-- MODAL CERTIFICADOS DE PUREZA -->
<div class="modal" id="certs">
  <div class="modal-panel">
    <button class="modal-close" onclick="closeModal('certs')">×</button>
    <h2>Certificados de pureza</h2>
    <p style="color:var(--muted);font-size:.88rem;margin-top:6px">Laudos de análise (HPLC) dos nossos lotes.</p>
    <div class="cert-list">
      <a href="certificados/1.pdf" target="_blank" rel="noopener">🔬 Certificado 1</a>
      <a href="certificados/2.pdf" target="_blank" rel="noopener">🔬 Certificado 2</a>
      <a href="certificados/3.pdf" target="_blank" rel="noopener">🔬 Certificado 3</a>
      <a href="certificados/4.pdf" target="_blank" rel="noopener">🔬 Certificado 4</a>
      <a href="certificados/5.pdf" target="_blank" rel="noopener">🔬 Certificado 5</a>
      <a href="certificados/6.pdf" target="_blank" rel="noopener">🔬 Certificado 6</a>
    </div>
  </div>
</div>

<!-- MODAL SHIPPING REQUIRED (aparece se checkout sem CEP) -->
<div class="modal" id="shipReq">
  <div class="modal-panel">
    <button class="modal-close" onclick="closeShipReq()">×</button>
    <div class="icon">📦</div>
    <h2>CEP obrigatório</h2>
    <p>Para calcular o frete e finalizar seu pedido, precisamos do seu CEP de entrega. Informe agora ou volte ao carrinho.</p>
    <input id="cepReq" placeholder="00000-000" maxlength="9"/>
    <div class="btns">
      <button class="btn-back" onclick="closeShipReq()">✕ Voltar ao carrinho</button>
      <button class="btn-cep" onclick="applyCepReq()">Calcular frete</button>
    </div>
  </div>
</div>

<style>
#ident select{width:100%;margin:8px 0;padding:13px 14px;border-radius:12px;border:1px solid #333;background:#0d0d0d;color:#fff;font-size:15px}
#ident .modal-panel{max-height:86vh;overflow:auto}
#ident input{width:100%;margin:8px 0;padding:13px 14px;border-radius:12px;border:1px solid #333;background:#0d0d0d;color:#fff;font-size:15px}
#ident .ident-err{color:#ff5252;font-size:13px;min-height:18px;margin:2px 0 8px}
</style>
<script>
/* =====================================================================
   DADOS INJETADOS PELO PYTHON
   ===================================================================== */
const PRODUCTS       = __PRODUCTS__;
const CATEGORY_COLORS= __CATCOLORS__;
const COUPONS        = __CUPONS__;
const REGIOES        = __REGIOES__;
const FRETES_CIDADES = __FRETES_CIDADES__;
const WHATSAPP_NUM   = "__WA__";
const FRETE_GRATIS   = __FRETE_GRATIS__;
const BRINDE_LIMITE  = __BRINDE_LIMITE__;
const ANALYTICS_URL  = "__ANALYTICS__";
const BACT_WATER     = PRODUCTS.find(p => p.nome === "BACTERIOSTATIC WATER");

/* =====================================================================
   ESTADO
   ===================================================================== */
let cart = {};             // {id: qty}
let filterCat = "TODOS";
let searchQ = "";
let orbitAngle = 0;
let orbitCount = 14;
let freteInfo = null;      // {uf, valor, cidade}

/* =====================================================================
   HELPERS
   ===================================================================== */
const brl = v => "R$ " + Number(v).toFixed(2).replace(".",",");
const $ = id => document.getElementById(id);
const openModal = id => $(id).classList.add("open");
const closeModal = id => $(id).classList.remove("open");
const closeShipReq = () => { closeModal("shipReq"); openModal("cart"); };

/* =====================================================================
   ORBITAL 3D — usa IMAGENS reais
   ===================================================================== */
function baseNome(nome){
  return String(nome || "")
    .normalize("NFD").replace(/[\u0300-\u036f]/g, "")
    .toUpperCase()
    .replace(/\s*\d+([.,]\d+)?\s*(MG|MCG|ML|UI|UN|UNI|G)\s*$/gi, "")
    .replace(/[^A-Z0-9]+/g, " ")
    .trim();
}
function buildOrbit(){
  const ring = $("orbitRing");
  // Todos os produtos disponíveis, sem repetir o mesmo peptídeo em outra dosagem
  const vistos = new Set();
  const items = PRODUCTS.filter(p => {
    if (!p.available) return false;
    const k = baseNome(p.nome);
    if (vistos.has(k)) return false;
    vistos.add(k);
    return true;
  });
  orbitCount = items.length || 1;
  const isMob = window.innerWidth < 768;
  const cardW = isMob ? 130 : 160;
  const gap   = isMob ? 26 : 34;
  // Raio calculado para que os cartões NUNCA se sobreponham, seja qual for a quantidade
  const minR  = (cardW + gap) / (2 * Math.tan(Math.PI / Math.max(orbitCount, 3)));
  const R = Math.max(isMob ? 300 : 420, Math.round(minR));
  const step = 360 / orbitCount;

  ring.innerHTML = items.map((p, i) => {
    const c = CATEGORY_COLORS[p.cat] || "#e10600";
    const img = p.img
      ? `<img src="${p.img}" alt="${p.nome}" onerror="this.style.display='none';this.nextElementSibling.style.display='block'"/><div class="fallback" style="display:none">${p.icon}</div>`
      : `<div class="fallback">${p.icon}</div>`;
    return `<div class="orbit-item" style="transform:rotateY(${i*step}deg) translateZ(${R}px)" onclick="openDetail(${p.id})">
      <div class="orbit-can" style="--can-color:${c}">
        <span class="tag">${p.cat}</span>
        ${img}
        <span class="name">${p.nome}</span>
      </div>
    </div>`;
  }).join("");
  // Ajusta perspectiva/zoom para que o anel inteiro caiba na tela
  const stage = $("orbitStage");
  if (stage) {
    stage.style.perspective = Math.round(R * 3.2) + "px";
    const fit = Math.min(1, (window.innerWidth * (isMob ? 0.95 : 0.9)) / (R * 2.1));
    stage.style.setProperty("--orbit-scale", fit.toFixed(3));
  }
  ring.style.transform = "rotateY(0deg)";

}
function rotOrbit(dir){
  const ring = $("orbitRing");
  ring.classList.add("paused");
  orbitAngle += dir * (360 / (orbitCount || 14));
  ring.style.animation = "none";
  ring.style.transform = `rotateY(${orbitAngle}deg)`;
  clearTimeout(window.__orbT);
  window.__orbT = setTimeout(() => {
    ring.style.animation = "spin 30s linear infinite";
    ring.classList.remove("paused");
  }, 3000);
}

/* =====================================================================
   FILTROS + GRID
   ===================================================================== */
function buildFilters(){
  const cats = ["TODOS", ...Object.keys(CATEGORY_COLORS)];
  $("filters").innerHTML =
    `<input class="search" placeholder="🔍 Buscar peptídeo..." oninput="searchQ=this.value.toLowerCase();renderGrid()"/>` +
    cats.map(c => `<button class="cat-pill ${c===filterCat?'active':''}" onclick="filterCat='${c}';buildFilters();renderGrid()">${c}</button>`).join("");
}
function renderGrid(){
  const list = PRODUCTS.filter(p => {
    const okCat = filterCat === "TODOS" || p.cat === filterCat;
    const okQ = !searchQ || p.nome.toLowerCase().includes(searchQ);
    return okCat && okQ;
  });
  $("grid").innerHTML = list.map(p => {
    const c = CATEGORY_COLORS[p.cat] || "#e10600";
    const img = p.img
      ? `<img src="${p.img}" alt="${p.nome}" onerror="this.style.display='none';this.nextElementSibling.style.display='block'"/><div class="fallback" style="display:none">${p.icon}</div>`
      : `<div class="fallback">${p.icon}</div>`;
    return `<div class="pcard ${!p.available?'unavailable':''}" style="--card-glow:${c}">
      <span class="pcard-tag">${p.cat}</span>
      <div class="pcard-img" onclick="openDetail(${p.id})">${img}</div>
      <div>
        <div class="pcard-name"><span class="pcard-emoji">${p.icon}</span> ${p.nome}</div>
        <div class="pcard-spec">${p.espec}</div>
      </div>
      <div class="pcard-price">
        <span class="cur">${brl(p.preco)}</span>
        ${p.promoPct>0?`<span class="old">${brl(p.precoOrig)}</span>`:""}
      </div>
      <button class="pcard-btn" type="button" data-action="add" data-id="${p.id}" ${!p.available?"disabled":""}>
        ${p.available ? "+ Adicionar" : "Indisponível"}
      </button>
    </div>`;
  }).join("");
}

/* =====================================================================
   DETALHE
   ===================================================================== */
function openDetail(id){
  const p = PRODUCTS.find(x => x.id === id);
  if (!p) return;
  const c = CATEGORY_COLORS[p.cat] || "#e10600";
  const img = p.img
    ? `<img src="${p.img}" alt="${p.nome}"/>`
    : `<div class="fallback">${p.icon}</div>`;
  $("detailBody").innerHTML = `
    <div class="detail-hero">
      <div class="detail-img" style="border-color:${c}">${img}</div>
      <div class="detail-info">
        <span class="pcard-tag" style="position:static;display:inline-block;margin-bottom:8px;border-color:${c}">${p.cat}</span>
        <h2>${p.nome}</h2>
        <div class="pcard-spec">${p.espec}</div>
        <div class="detail-price">
          <span class="cur">${brl(p.preco)}</span>
          ${p.promoPct>0?`<span class="old">${brl(p.precoOrig)}</span>`:""}
        </div>
        <button class="pcard-btn" style="padding:12px 24px" type="button" data-action="add-close" data-id="${p.id}" ${!p.available?"disabled":""}>
          ${p.available ? "+ Adicionar ao carrinho" : "Indisponível"}
        </button>
      </div>
    </div>
    <div class="body-txt">${p.info}</div>`;
  openModal("detail");
}

/* =====================================================================
   CARRINHO + BRINDE
   ===================================================================== */
function toast(msg){
  let t = $("toast");
  if (!t){
    t = document.createElement("div");
    t.id = "toast"; t.className = "toast";
    document.body.appendChild(t);
  }
  t.textContent = msg;
  t.classList.add("show");
  clearTimeout(t._tid);
  t._tid = setTimeout(() => t.classList.remove("show"), 2200);
}
function addCart(id){
  const p = PRODUCTS.find(x => x.id === Number(id));
  if (!p){ console.warn("produto inexistente", id); return; }
  if (!p.available){ toast("Produto indisponivel no momento."); return; }
  cart[p.id] = (cart[p.id] || 0) + 1;
  updateCart();
  ["cartBtn","cartFab"].forEach(function(bid){
    const btn = $(bid);
    if (btn){ btn.classList.remove("pulse"); void btn.offsetWidth; btn.classList.add("pulse"); }
  });
  toast(`\u2713 ${p.nome} adicionado ao carrinho (${cart[p.id]}\u00d7)`);
  track(__trackouAdd ? "add_carrinho" : "primeiro_add_carrinho", { produto: `${p.nome} ${p.espec}` });
  __trackouAdd = true;
  window.__abandonoEnviado = false;
}
function rmCart(id){ delete cart[id]; updateCart(); }
function qtyCart(id, d){
  cart[id] = Math.max(0, (cart[id] || 0) + d);
  if (cart[id] === 0) delete cart[id];
  updateCart();
}
function cartSubtotal(){
  return Object.entries(cart).reduce((s,[id,q]) => {
    const p = PRODUCTS.find(x => x.id === +id);
    return s + (p ? p.preco * q : 0);
  }, 0);
}
function cartUnitsTotal(){
  return Object.values(cart).reduce((s,q) => s+q, 0);
}
function isBrinde(){
  return cartSubtotal() >= BRINDE_LIMITE && cartUnitsTotal() > 0;
}
function brindeQtd(){
  return isBrinde() ? cartUnitsTotal() : 0;
}
function updateCart(){
  const n = cartUnitsTotal();
  const c = $("cartCount"); if (c) c.textContent = n;
  const fc = $("cartFabCount"); if (fc) fc.textContent = n;
  const fab = $("cartFab"); if (fab) fab.classList.toggle("empty", n === 0);
  renderCart();
}
function renderCart(){
  const items = Object.entries(cart).map(([id, q]) => {
    const p = PRODUCTS.find(x => x.id === +id);
    if (!p) return "";
    const img = p.img ? `<img src="${p.img}" onerror="this.style.display='none'"/>` : `<div style="width:44px;height:44px;background:#000;border-radius:8px;display:grid;place-items:center">${p.icon}</div>`;
    return `<div class="cart-item">
      ${img}
      <div class="info">
        <div class="name">${p.nome} — ${p.espec}</div>
        <div class="price">${brl(p.preco * q)}</div>
      </div>
      <div class="qty">
        <button onclick="qtyCart(${p.id},-1)">−</button>
        <span>${q}</span>
        <button onclick="qtyCart(${p.id},1)">+</button>
      </div>
      <button class="rm" onclick="rmCart(${p.id})">✕</button>
    </div>`;
  }).join("");

  // BRINDE
  let brindeHtml = "";
  const bq = brindeQtd();
  if (bq > 0 && BACT_WATER){
    brindeHtml = `<div class="cart-item brinde">
      ${BACT_WATER.img ? `<img src="${BACT_WATER.img}" onerror="this.style.display='none'"/>` : `<div style="width:44px;height:44px;background:#000;border-radius:8px;display:grid;place-items:center">💧</div>`}
      <div class="info">
        <div class="name">${BACT_WATER.nome} — ${BACT_WATER.espec} <span class="tag-brinde">BRINDE</span></div>
        <div class="price">GRÁTIS × ${bq}</div>
      </div>
    </div>`;
  }
  $("cartItems").innerHTML = (items || `<p class="subtle">Carrinho vazio.</p>`) + brindeHtml;

  // aviso brinde
  const sub = cartSubtotal();
  if (sub > 0 && sub < BRINDE_LIMITE){
    const falta = BRINDE_LIMITE - sub;
    $("brindeMsg").innerHTML = `<div class="brinde-warn">🎁 Faltam <b>${brl(falta)}</b> para ganhar <b>Bacteriostatic Water GRÁTIS</b> em cada item do seu pedido!</div>`;
  } else if (isBrinde()){
    $("brindeMsg").innerHTML = `<div class="brinde-warn">🎁 <b>PARABÉNS!</b> Você ganhou <b>${bq}× Bacteriostatic Water</b> como BRINDE!</div>`;
  } else {
    $("brindeMsg").innerHTML = "";
  }

  calcTotals();
}

/* =====================================================================
   CEP / FRETE
   ===================================================================== */
function ufToRegiao(uf){
  for (const [nome, r] of Object.entries(REGIOES)){
    if (r.ufs.includes(uf)) return {regiao: nome, frete: r.frete, prazo: r.prazo || ""};
  }
  return null;
}
/* Consulta CEP em DUAS fontes: ViaCEP e BrasilAPI (fallback automatico) */
async function buscaViaCep(cep){
  const r = await fetch(`https://viacep.com.br/ws/${cep}/json/`);
  if (!r.ok) throw new Error("viacep http " + r.status);
  const d = await r.json();
  if (d.erro) throw new Error("cep nao encontrado (viacep)");
  return { uf: d.uf, cidade: d.localidade, bairro: d.bairro || "", logradouro: d.logradouro || "", fonte: "ViaCEP" };
}
async function buscaBrasilApi(cep){
  const r = await fetch(`https://brasilapi.com.br/api/cep/v1/${cep}`);
  if (!r.ok) throw new Error("brasilapi http " + r.status);
  const d = await r.json();
  if (!d.state) throw new Error("cep nao encontrado (brasilapi)");
  return { uf: d.state, cidade: d.city, bairro: d.neighborhood || "", logradouro: d.street || "", fonte: "BrasilAPI" };
}
async function lookupCep(cep){
  const fontes = [buscaViaCep, buscaBrasilApi];
  let ultimoErro = null;
  for (const f of fontes){
    try { return await f(cep); }
    catch(e){ ultimoErro = e; }
  }
  throw ultimoErro || new Error("CEP nao localizado");
}
let cepTimer = null;
function onCepInput(){
  clearTimeout(cepTimer);
  cepTimer = setTimeout(calcFrete, 400);
}
async function calcFrete(){
  const el = $("cepInput");
  const cep = (el.value || "").replace(/\D/g,"").slice(0,8);
  const st = $("cepStatus");
  if (cep.length !== 8){
    freteInfo = null;
    if (st) st.textContent = cep.length ? "Digite os 8 digitos do CEP." : "";
    calcTotals();
    return;
  }
  if (st) st.textContent = "Consultando CEP...";
  try{
    const d = await lookupCep(cep);
    const uf = (d.uf || "").toUpperCase().replace(/[^A-Z]/g,"").slice(0,2);
    const cidadeNorm = (d.cidade || "").normalize("NFD").replace(/[\u0300-\u036f]/g,"").toUpperCase();
    const chaveCidade = cidadeNorm + "-" + uf;
    let valor, regiao, prazo;
    if (FRETES_CIDADES[chaveCidade] !== undefined){
      valor = FRETES_CIDADES[chaveCidade]; regiao = "REGIONAL"; prazo = "1-3 dias";
    } else {
      const info = ufToRegiao(uf);
      if (!info) throw new Error("UF nao atendida");
      valor = info.frete; regiao = info.regiao; prazo = info.prazo || "";
    }
    freteInfo = {uf, cidade: d.cidade, bairro: d.bairro, valor, regiao, prazo, fonte: d.fonte};
    preencherEndereco(d, cep);
    if (st) st.textContent = `\u2713 ${d.cidade}/${uf} \u2014 ${regiao} ${brl(valor)}${prazo ? " (" + prazo + ")" : ""}`;
  }catch(e){
    freteInfo = null;
    if (st) st.textContent = "\u26a0 CEP nao encontrado. Confira e tente novamente.";
  }
  calcTotals();
  if (freteInfo) track("cep_informado");
}
function calcTotals(){
  const sub = cartSubtotal();
  const cup = ($("cupomInput").value || "").trim().toUpperCase();
  const desc = COUPONS[cup] ? sub * COUPONS[cup] : 0;
  const freteBruto = freteInfo ? freteInfo.valor : 0;
  const frete = (sub >= FRETE_GRATIS && freteInfo) ? 0 : freteBruto;
  const total = sub - desc + frete;
  $("totals").innerHTML = `
    <div class="row"><span>Subtotal</span><span>${brl(sub)}</span></div>
    ${desc>0?`<div class="row" style="color:var(--red)"><span>Cupom ${cup}</span><span>− ${brl(desc)}</span></div>`:""}
    ${freteInfo?`<div class="row"><span>Frete ${freteInfo.regiao} (${freteInfo.cidade}/${freteInfo.uf})</span><span>${frete===0?"GRÁTIS":brl(frete)}</span></div>`:`<div class="row" style="color:var(--muted)"><span>Frete</span><span>informe o CEP</span></div>`}
    ${isBrinde()?`<div class="row" style="color:var(--red)"><span>🎁 Brinde (${brindeQtd()}× Bact. Water)</span><span>GRÁTIS</span></div>`:""}
    <div class="row tot"><span>TOTAL</span><span>${brl(total)}</span></div>`;
}

/* =====================================================================
   RASTREIO DE CLIENTES (Google Sheets / Apps Script)
   Eventos: visita | add_carrinho | abriu_carrinho | cep | identificado
            | checkout_whatsapp | carrinho_abandonado
   ===================================================================== */
const LS = {
  get(k, d){ try { return JSON.parse(localStorage.getItem(k)) ?? d; } catch(e){ return d; } },
  set(k, v){ try { localStorage.setItem(k, JSON.stringify(v)); } catch(e){} }
};
function visitorId(){
  let v = LS.get("glab_vid", null);
  if (!v){
    v = "V-" + Date.now().toString(36) + "-" + Math.random().toString(36).slice(2,8);
    LS.set("glab_vid", v);
  }
  return v;
}
function sessionId(){
  if (!window.__sid) window.__sid = "S-" + Date.now().toString(36) + "-" + Math.random().toString(36).slice(2,6);
  return window.__sid;
}
function cartResumo(){
  return Object.entries(cart).map(([id,q]) => {
    const p = PRODUCTS.find(x => x.id === +id);
    return p ? `${q}x ${p.nome} ${p.espec}` : "";
  }).filter(Boolean).join(" | ");
}
function track(evento, extra, useBeacon){
  if (!ANALYTICS_URL) return;
  const cli = LS.get("glab_cliente", {}) || {};
  const payload = Object.assign({
    evento,
    ts: new Date().toISOString(),
    visitante: visitorId(),
    sessao: sessionId(),
    nome: cli.nome || "",
    whatsapp: cli.zap || "",
    itens: cartResumo(),
    unidades: cartUnitsTotal(),
    subtotal: cartSubtotal().toFixed(2),
    cidade: freteInfo ? freteInfo.cidade : "",
    uf: freteInfo ? freteInfo.uf : "",
    regiao: freteInfo ? freteInfo.regiao : "",
    cupom: ($("cupomInput") && $("cupomInput").value || "").trim().toUpperCase(),
    url: location.href,
    userAgent: navigator.userAgent
  }, extra || {});
  const body = JSON.stringify(payload);
  try {
    if (useBeacon && navigator.sendBeacon){
      navigator.sendBeacon(ANALYTICS_URL, new Blob([body], {type: "text/plain;charset=UTF-8"}));
    } else {
      fetch(ANALYTICS_URL, { method: "POST", mode: "no-cors", keepalive: true,
        headers: {"Content-Type": "text/plain;charset=UTF-8"}, body });
    }
  } catch(e){ console.warn("[G-LAB] track falhou", e); }
}
/* carrinho abandonado: houve itens, mas nao houve envio ao WhatsApp */
let __enviouZap = false;
let __trackouAdd = false;
function registrarAbandono(){
  if (__enviouZap || cartUnitsTotal() === 0 || window.__abandonoEnviado) return;
  window.__abandonoEnviado = true;
  track("carrinho_abandonado", {}, true);
}
window.addEventListener("pagehide", registrarAbandono);
document.addEventListener("visibilitychange", () => {
  if (document.visibilityState === "hidden") registrarAbandono();
});

/* =====================================================================
   CHECKOUT (identificacao -> CEP -> WhatsApp)
   ===================================================================== */
function closeIdent(){ closeModal("ident"); openModal("cart"); }

/* sanitiza entradas do formulario */
function sanitizarEntrada(valor, max){
  return String(valor || "")
    .replace(/[<>]/g, "")
    .replace(/\s+/g, " ")
    .trim()
    .slice(0, max);
}
function mascaraCPF(el){
  el.value = el.value.replace(/\D/g,"").slice(0,11)
    .replace(/(\d{3})(\d)/, "$1.$2")
    .replace(/(\d{3})\.(\d{3})(\d)/, "$1.$2.$3")
    .replace(/(\d{3})\.(\d{3})\.(\d{3})(\d)/, "$1.$2.$3-$4");
}
function cpfValido(cpf){
  const c = String(cpf).replace(/\D/g,"");
  if (c.length !== 11 || /^(\d)\1{10}$/.test(c)) return false;
  let s1 = 0, s2 = 0;
  for (let i = 0; i < 9; i++) s1 += +c[i] * (10 - i);
  let d1 = (s1 * 10) % 11; if (d1 === 10) d1 = 0;
  if (d1 !== +c[9]) return false;
  for (let i = 0; i < 10; i++) s2 += +c[i] * (11 - i);
  let d2 = (s2 * 10) % 11; if (d2 === 10) d2 = 0;
  return d2 === +c[10];
}
function preencherEndereco(d, cep){
  const set = (id, v) => { const el = $(id); if (el && !el.value && v) el.value = v; };
  set("f_end", d.logradouro || "");
  set("f_bairro", d.bairro || "");
  set("f_cidade", d.cidade || "");
  set("f_estado", (d.uf || "").toUpperCase());
  const c = $("cep-destino");
  if (c && cep) c.value = cep.replace(/(\d{5})(\d{3})/, "$1-$2");
}
function confirmIdent(){ enviarPedido(); }
function checkout(){
  if (cartUnitsTotal() === 0){ alert("Adicione produtos primeiro!"); return; }
  if (!freteInfo){
    closeModal("cart");
    openModal("shipReq");
    return;
  }
  const cli = LS.get("glab_cliente", {}) || {};
  closeModal("cart");
  ["f_nome","f_cpf","f_tel","f_end","f_num","f_comp","f_bairro","f_cidade","f_estado","cep-destino"].forEach(id => {
    const el = $(id);
    if (el && !el.value && cli[id]) el.value = cli[id];
  });
  if ($("f_cidade") && !$("f_cidade").value) $("f_cidade").value = freteInfo.cidade || "";
  if ($("f_estado") && !$("f_estado").value) $("f_estado").value = freteInfo.uf || "";
  if ($("f_bairro") && !$("f_bairro").value) $("f_bairro").value = freteInfo.bairro || "";
  if ($("cep-destino") && !$("cep-destino").value) $("cep-destino").value = ($("cepInput").value || "");
  openModal("ident");
}
function enviarPedido(){
  const d = {
    n:   sanitizarEntrada($("f_nome").value,   120).toUpperCase(),
    cpf: sanitizarEntrada($("f_cpf").value,     14),
    e:   sanitizarEntrada($("f_end").value,    200).toUpperCase(),
    nu:  sanitizarEntrada($("f_num").value,     10),
    ba:  sanitizarEntrada($("f_bairro").value, 100).toUpperCase(),
    co:  sanitizarEntrada($("f_comp").value,   100).toUpperCase(),
    ci:  sanitizarEntrada($("f_cidade").value, 100).toUpperCase(),
    es:  sanitizarEntrada($("f_estado").value,   2).toUpperCase(),
    ce:  $("cep-destino").value.replace(/\D/g,"").replace(/(\d{5})(\d{3})/, "$1-$2"),
    t:   sanitizarEntrada($("f_tel").value,     20),
    p:   $("f_pgto").value === "Pix" ? "PIX" : "CART\u00c3O DE CR\u00c9DITO",
    cupom: ($("cupomInput").value || "").trim().toUpperCase()
  };
  const err = $("identErr");
  const falhar = m => { if (err) err.textContent = m; else alert(m); };
  if (!d.n || !d.cpf || !d.e || !d.nu || !d.ba || !d.ci || !d.es || !d.t){
    falhar("Preencha todos os campos obrigat\u00f3rios!");
    return;
  }
  if (!cpfValido(d.cpf)){ falhar("CPF inv\u00e1lido."); return; }
  if (d.t.replace(/\D/g,"").length < 10){ falhar("WhatsApp inv\u00e1lido (use DDD + n\u00famero)."); return; }
  if (d.ce.replace(/\D/g,"").length !== 8){ falhar("CEP inv\u00e1lido."); return; }
  if (err) err.textContent = "";

  LS.set("glab_cliente", {
    nome: d.n, zap: d.t,
    f_nome: d.n, f_cpf: d.cpf, f_tel: d.t, f_end: d.e, f_num: d.nu,
    f_comp: d.co, f_bairro: d.ba, f_cidade: d.ci, f_estado: d.es, "cep-destino": d.ce
  });
  track("identificado");
  closeModal("ident");

  const sub = cartSubtotal();
  const cup = ($("cupomInput").value || "").trim().toUpperCase();
  const desc = COUPONS[cup] ? sub * COUPONS[cup] : 0;
  const frete = sub >= FRETE_GRATIS ? 0 : freteInfo.valor;
  const total = sub - desc + frete;

  const pct = COUPONS[cup] || 0;
  let msg = `*NOVO PEDIDO G-LAB*\n\n`;
  msg += `*CLIENTE:*\n`;
  msg += `\u2022 *NOME:* ${d.n}\n`;
  msg += `\u2022 *CPF:* ${d.cpf}\n`;
  msg += `\u2022 *WHATSAPP:* ${d.t}\n`;
  msg += `\u2022 *END:* ${d.e} - N${d.nu}\n`;
  msg += `\u2022 *BAIRRO:* ${d.ba}\n`;
  if (d.co) msg += `\u2022 *COMPL:* ${d.co}\n`;
  msg += `\u2022 *CIDADE:* ${d.ci}-${d.es}\n`;
  msg += `\u2022 *CEP:* ${d.ce}\n`;
  msg += `\u2022 *PGTO:* ${d.p}\n`;
  msg += `\n*cupom:* ${cup || "NENHUM"}\n`;
  msg += `\n*ITENS:*\n`;
  for (const [id, q] of Object.entries(cart)){
    const p = PRODUCTS.find(x => x.id === +id);
    const bruto = p.preco * q;
    msg += `\u2022 ${q}x ${p.nome} (${p.espec}) - ${brl(bruto)}`;
    if (pct > 0) msg += ` - COM DESCONTO ${brl(bruto * (1 - pct))}`;
    msg += `\n`;
  }
  const bq = brindeQtd();
  if (bq > 0){
    msg += `\u2022 ${bq}x ${BACT_WATER.nome} ${BACT_WATER.espec} - BRINDE\n`;
  }
  msg += `\n\ud83d\ude9a *FRETE:* ${frete === 0 ? "GR\u00c1TIS" : (freteInfo.regiao + " " + brl(frete) + (freteInfo.prazo ? " (" + freteInfo.prazo.toUpperCase() + ")" : ""))}\n`;
  msg += `\n*TOTAL: ${brl(total)}*`;


  __enviouZap = true;
  track("checkout_whatsapp", { total: total.toFixed(2), frete: frete.toFixed(2), desconto: desc.toFixed(2), brinde: bq });
  window.open(`https://wa.me/${WHATSAPP_NUM}?text=${encodeURIComponent(msg)}`, "_blank");
}
function applyCepReq(){
  $("cepInput").value = $("cepReq").value;
  const cd = $("cep-destino"); if (cd) cd.value = $("cepReq").value;
  calcFrete().then(() => {
    closeModal("shipReq");
    openModal("cart");
  });
}

/* =====================================================================
   INIT
   ===================================================================== */
buildFilters();
renderGrid();
buildOrbit();
updateCart();
track("visita");
document.addEventListener("click", (ev) => {
  const el = ev.target.closest("[data-action]");
  if (!el) return;
  const act = el.dataset.action;
  const id  = el.dataset.id;
  if (act === "add"){ ev.preventDefault(); addCart(id); }
  else if (act === "add-close"){ ev.preventDefault(); addCart(id); closeModal("detail"); }
  else if (act === "open-cart"){ ev.preventDefault(); openModal("cart"); track("abriu_carrinho"); }
});
window.addEventListener("resize", () => buildOrbit());
window.addEventListener("error", (e) => console.error("[G-LAB]", e.message));

/* =====================================================================
   CALCULADORA DE PEPTÍDEOS
   ===================================================================== */
const CALC_DOSES_MCG = [100,150,200,250,300,400,500,600,750,1000];
const CALC_DOSES_MG  = [1,1.5,2,2.5,3,4,5,7.5,10];
const CALC_VIALS     = [2,5,10,15,20,30,50,100,500];
const CALC_DILS      = [0.5,1,1.5,2,2.5,3,4,5];
const CALC_SYRS      = [{ml:0.3,u:30,lab:"0,3 ml (30 UI)"},{ml:0.5,u:50,lab:"0,5 ml (50 UI)"},{ml:1,u:100,lab:"1 ml (100 UI)"}];

const calcState = { unit:"mcg", dose:250, vial:10, dil:2, syr:1 };

function calcNum(v){ const n = parseFloat(String(v).replace(",", ".")); return isFinite(n) && n > 0 ? n : null; }
function calcFmt(n, d){
  return n.toLocaleString("pt-BR", { minimumFractionDigits:0, maximumFractionDigits:(d===undefined?2:d) });
}
function calcSetUnit(u){
  if (calcState.unit === u) return;
  const d = calcState.dose;
  calcState.unit = u;
  if (d != null) calcState.dose = (u === "mg") ? d/1000 : d*1000;
  $("calcDoseCustom").value = "";
  calcRender();
}
function calcPick(key, val){ calcState[key] = val; 
  const map = { dose:"calcDoseCustom", vial:"calcVialCustom", dil:"calcDilCustom" };
  if (map[key]) $(map[key]).value = "";
  calcRender();
}
function calcReset(){
  Object.assign(calcState, { unit:"mcg", dose:250, vial:10, dil:2, syr:1 });
  ["calcDoseCustom","calcVialCustom","calcDilCustom"].forEach(i => { const e = $(i); if (e) e.value = ""; });
  calcRender();
}
function calcOptsHtml(list, key, fmt){
  return list.map(v => `<button type="button" class="calc-opt${calcState[key] === v ? " active":""}"
    onclick="calcPick('${key}',${v})">${fmt(v)}</button>`).join("");
}
function calcSyringeSvg(maxU, u){
  const W = 520, H = 96, x0 = 46, x1 = 470, yT = 30, hB = 30;
  const frac = maxU > 0 ? Math.max(0, Math.min(1, u / maxU)) : 0;
  const fw = (x1 - x0) * frac;
  const step = 1;
  let ticks = "";
  for (let t = 0; t <= maxU + 0.001; t += step){
    const x = x0 + (x1 - x0) * (t / maxU);
    const big = (t % 5 === 0) || t === maxU;
    ticks += `<line x1="${x.toFixed(1)}" y1="${yT}" x2="${x.toFixed(1)}" y2="${(yT + (big?11:7)).toFixed(1)}"
      stroke="rgba(255,255,255,${big?0.85:0.45})" stroke-width="${big?1.6:1}"/>`;
    if (big) ticks += `<text x="${x.toFixed(1)}" y="${yT-6}" fill="rgba(255,255,255,.7)" font-size="12"
      font-family="monospace" text-anchor="middle">${t}</text>`;
  }
  const mx = x0 + (x1 - x0) * frac;
  const marker = u > 0 ? `
    <line x1="${mx.toFixed(1)}" y1="${yT-16}" x2="${mx.toFixed(1)}" y2="${yT+hB+10}" stroke="#e10600" stroke-width="2"/>
    <circle cx="${mx.toFixed(1)}" cy="${yT+hB+14}" r="3.4" fill="#e10600"/>
    <text x="${Math.min(mx, x1-26).toFixed(1)}" y="${H-4}" fill="#ff6b66" font-size="12" font-weight="700"
      font-family="monospace" text-anchor="middle">${calcFmt(u,1)} UI</text>` : "";
  return `<svg viewBox="0 0 ${W} ${H}" role="img" aria-label="Seringa de insulina">
    <rect x="${x0-34}" y="${yT+hB/2-3}" width="34" height="6" rx="2" fill="rgba(255,255,255,.35)"/>
    <rect x="${x0-52}" y="${yT+hB/2-1.2}" width="20" height="2.4" fill="rgba(255,255,255,.55)"/>
    <rect x="${x0}" y="${yT}" width="${x1-x0}" height="${hB}" rx="6" fill="rgba(255,255,255,.06)"
      stroke="rgba(255,255,255,.28)"/>
    <rect x="${x0}" y="${yT}" width="${fw.toFixed(1)}" height="${hB}" rx="6" fill="rgba(225,6,0,.55)"/>
    <rect x="${(x0+fw).toFixed(1)}" y="${yT-4}" width="7" height="${hB+8}" rx="2" fill="rgba(255,255,255,.75)"/>
    <rect x="${x1}" y="${yT-6}" width="8" height="${hB+12}" rx="3" fill="rgba(255,255,255,.3)"/>
    <rect x="${x1+8}" y="${yT+hB/2-2.5}" width="34" height="5" rx="2" fill="rgba(255,255,255,.35)"/>
    ${ticks}${marker}
  </svg>`;
}
function calcRender(){
  const st = calcState;
  const doses = st.unit === "mg" ? CALC_DOSES_MG : CALC_DOSES_MCG;
  $("calcDoseOpts").innerHTML = calcOptsHtml(doses, "dose", v => `${calcFmt(v,2)} ${st.unit}`);
  $("calcVialOpts").innerHTML = calcOptsHtml(CALC_VIALS, "vial", v => `${calcFmt(v,2)} mg`);
  $("calcDilOpts").innerHTML  = calcOptsHtml(CALC_DILS, "dil", v => `${calcFmt(v,2)} ml`);
  $("calcSyrOpts").innerHTML  = CALC_SYRS.map(s => `<button type="button"
    class="calc-opt${st.syr === s.ml ? " active":""}" onclick="calcPick('syr',${s.ml})">${s.lab}</button>`).join("");
  $("calcDoseUnit").textContent = st.unit;
  $("calcUnitMcg").classList.toggle("active", st.unit === "mcg");
  $("calcUnitMg").classList.toggle("active", st.unit === "mg");

  const doseMg = st.dose == null ? null : (st.unit === "mg" ? st.dose : st.dose/1000);
  const vial = st.vial, dil = st.dil;
  const syr = CALC_SYRS.find(s => s.ml === st.syr) || CALC_SYRS[2];
  const ok = doseMg && vial && dil;
  if (!ok){
    $("calcUI").textContent = "—"; $("calcML").textContent = "—"; $("calcConc").textContent = "—";
    $("calcSyr").innerHTML = calcSyringeSvg(syr.u, 0);
    $("calcExtra").innerHTML = "Escolha a dose, o frasco e o diluente.";
    $("calcWarn").innerHTML = "";
    return;
  }
  const conc = vial / dil;               // mg/ml
  const ml   = doseMg / conc;            // ml por aplicação
  const ui   = ml * 100;                 // UI (seringa U-100)
  const aplic = Math.floor(vial / doseMg);

  $("calcConc").innerHTML = `${calcFmt(conc,2)} <em>mg/ml</em>`;
  $("calcML").innerHTML   = `${calcFmt(ml,3)} <em>ml</em>`;
  $("calcUI").innerHTML   = `${calcFmt(ui,1)} <em>UI</em>`;
  $("calcSyr").innerHTML  = calcSyringeSvg(syr.u, ui);
  $("calcExtra").innerHTML = `Frasco de <b>${calcFmt(vial,2)} mg</b> diluído em <b>${calcFmt(dil,2)} ml</b> → cada
    <b>${calcFmt(ui,1)} UI</b> (${calcFmt(ml,3)} ml) entrega <b>${st.unit === "mg" ? calcFmt(doseMg,3)+" mg" : calcFmt(doseMg*1000,0)+" mcg"}</b>.
    Rende aproximadamente <b>${aplic}</b> aplicaç${aplic === 1 ? "ão" : "ões"} por frasco.`;
  let w = "";
  if (ui > syr.u) w = `⚠ A dose ocupa ${calcFmt(ui,1)} UI e não cabe na seringa de ${syr.lab}. Use uma seringa maior, divida em 2 aplicações ou reduza o diluente.`;
  else if (ui < 2) w = `⚠ Volume muito pequeno (${calcFmt(ui,1)} UI) — difícil de medir com precisão. Use menos diluente.`;
  $("calcWarn").innerHTML = w ? `<div class="calc-warn">${w}</div>` : "";

  const zap = $("calcZap");
  if (zap) zap.href = `https://wa.me/${WHATSAPP_NUM}?text=` + encodeURIComponent(
    `Olá! Dúvida sobre reconstituição: frasco de ${calcFmt(vial,2)} mg, ${calcFmt(dil,2)} ml de diluente, dose de ${st.unit === "mg" ? calcFmt(doseMg,3)+" mg" : calcFmt(doseMg*1000,0)+" mcg"} (${calcFmt(ui,1)} UI).`);
}
["calcDoseCustom","calcVialCustom","calcDilCustom"].forEach(id => {
  const el = $(id); if (!el) return;
  el.addEventListener("input", () => {
    const key = id === "calcDoseCustom" ? "dose" : (id === "calcVialCustom" ? "vial" : "dil");
    calcState[key] = calcNum(el.value);
    calcRender();
  });
});
calcRender();
</script>
</body>
</html>
"""

# =====================================================================
# 8) GERAÇÃO
# =====================================================================
def _forcar_remocao(func, path, _exc):
    """Tenta remover arquivo/pasta somente-leitura (Windows/OneDrive)."""
    try:
        os.chmod(path, stat.S_IWRITE)
        func(path)
    except Exception:
        pass


def sincronizar_imagens(origem: Path, destino: Path) -> int:
    """
    Copia as imagens SEM apagar a pasta de destino.
    OneDrive / antivírus / Explorer costumam manter um handle aberto na pasta,
    o que fazia o shutil.rmtree() falhar com PermissionError [WinError 5].
    Aqui só sobrescrevemos arquivo por arquivo e limpamos os obsoletos.
    """
    destino.mkdir(parents=True, exist_ok=True)
    copiados = 0
    nomes_origem = set()

    for src in origem.iterdir():
        if not src.is_file():
            continue
        nomes_origem.add(src.name)
        dst = destino / src.name
        try:
            # pula se já está idêntico (mesmo tamanho e mais novo)
            if dst.exists() and dst.stat().st_size == src.stat().st_size \
               and dst.stat().st_mtime >= src.stat().st_mtime:
                continue
            if dst.exists():
                try:
                    os.chmod(dst, stat.S_IWRITE)
                except Exception:
                    pass
            shutil.copy2(src, dst)
            copiados += 1
        except PermissionError:
            print(f"   ⚠ sem permissão para atualizar '{src.name}' (arquivo em uso) — mantido o antigo.")
        except Exception as e:
            print(f"   ⚠ falha ao copiar '{src.name}': {e}")

    # remove arquivos que não existem mais na origem (best-effort)
    for antigo in destino.iterdir():
        if antigo.is_file() and antigo.name not in nomes_origem:
            try:
                os.chmod(antigo, stat.S_IWRITE)
                antigo.unlink()
            except Exception:
                pass

    return copiados


def gerar():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # copia pasta de imagens se existir (sem rmtree — evita WinError 5)
    if IMG_DIR.exists():
        n = sincronizar_imagens(IMG_DIR, OUTPUT_IMG_DIR)
        total = len([p for p in OUTPUT_IMG_DIR.iterdir() if p.is_file()])
        print(f"✓  Imagens sincronizadas: {n} atualizadas / {total} no total.")
    else:
        print(f"⚠  Pasta '{IMG_DIR.name}' não encontrada — o site funcionará com ícones fallback.")

    print(f"\u2139  Planilha usada: {XLSX_PATH.name if XLSX_PATH.exists() else '(nenhuma encontrada)'}")
    overrides = carregar_excel(XLSX_PATH)
    produtos  = montar_produtos(overrides)

    html = (HTML_TEMPLATE
        .replace("__PRODUCTS__",     json.dumps(produtos, ensure_ascii=False))
        .replace("__CATCOLORS__",    json.dumps(CATEGORY_COLORS, ensure_ascii=False))
        .replace("__CUPONS__",       json.dumps(CUPONS))
        .replace("__REGIOES__",      json.dumps(REGIOES, ensure_ascii=False))
        .replace("__FRETES_CIDADES__", json.dumps(FRETES_CIDADES, ensure_ascii=False))
        .replace("__WA__",           WHATSAPP_NUM)
        .replace("__FRETE_GRATIS__", str(FRETE_GRATIS))
        .replace("__BRINDE_LIMITE__",str(BRINDE_LIMITE))
        .replace("__ANALYTICS__",    ANALYTICS_URL)
    )
    try:
        OUTPUT_HTML.write_text(html, encoding="utf-8")
    except PermissionError:
        alt = OUTPUT_DIR / "index_novo.html"
        alt.write_text(html, encoding="utf-8")
        print(f"⚠  '{OUTPUT_HTML.name}' está aberto/bloqueado. Salvei em: {alt}")
        return
    print(f"\n✅ Site gerado: {OUTPUT_HTML}")
    print(f"   Abra no navegador: file://{OUTPUT_HTML.resolve()}")

if __name__ == "__main__":
    gerar()
